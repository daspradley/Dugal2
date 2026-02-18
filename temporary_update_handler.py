"""
TemporaryUpdateHandler - All Write Operations
Handles ALL write operations via temporary local files.
Creates a temp working copy, makes changes, tracks modifications.
Works with UniversalFileHandler for read operations.
"""

import logging
import shutil
import tempfile
from pathlib import Path
from typing import Dict, Any, Optional, List
from openpyxl import load_workbook
from datetime import datetime
import json

logger = logging.getLogger(__name__)


class TemporaryUpdateHandler:
    """
    Handles all write operations via temporary local files.
    
    Workflow:
    1. Create temp copy of source file
    2. Make all updates to temp file
    3. Track changes made
    4. Temp file can be synced back to source (via SyncManager)
    
    Supports:
    - Excel files (.xlsx, .xlsm)
    - Google Sheets (download as Excel, update, re-upload)
    - CSV files
    """
    
    def __init__(self, source_path: str, file_type: str, file_handler=None):
        """
        Initialize TemporaryUpdateHandler.
        
        Args:
            source_path: Original file path/URL (OneDrive, Google Sheets, etc.)
            file_type: Type of file (from UniversalFileHandler constants)
            file_handler: Reference to UniversalFileHandler for read operations
        """
        self.source_path = source_path
        self.file_type = file_type
        self.file_handler = file_handler
        
        # Create temp directory
        self.temp_dir = Path(tempfile.gettempdir()) / "Dugal_Working"
        self.temp_dir.mkdir(exist_ok=True)
        
        # Generate temp file path
        self.temp_path = None
        self.workbook = None
        
        # Change tracking
        self.changes_made = False
        self.change_log = []
        self.last_save_time = None
        
        # Initialize temp file
        self._create_temp_file()
        
        logger.info(f"TemporaryUpdateHandler initialized: {self.temp_path}")
    
    def _create_temp_file(self):
        """Create temporary working copy of the source file."""
        try:
            # Generate temp filename
            if self.file_type == "google_sheets":
                # Google Sheets: Extract name from file_handler
                if self.file_handler and self.file_handler.file_info:
                    filename = self.file_handler.file_info.get('filename', 'sheet')
                else:
                    filename = "google_sheet"
                filename = f"{filename}_WORKING.xlsx"
            else:
                # Excel/CSV: Use original filename
                original_name = Path(self.source_path).stem
                filename = f"{original_name}_WORKING.xlsx"
            
            self.temp_path = self.temp_dir / filename
            
            # Copy source to temp
            if self.file_type in ["excel", "csv"]:
                # Local file - direct copy
                logger.debug(f"Copying {self.source_path} → {self.temp_path}")
                shutil.copy(self.source_path, self.temp_path)
            
            elif self.file_type == "google_sheets":
                # Google Sheets - download as Excel
                logger.debug(f"Downloading Google Sheet to temp: {self.temp_path}")
                self._download_google_sheet_as_excel()
            
            # Open temp file for writing
            self.workbook = load_workbook(self.temp_path, data_only=False)
            self.last_save_time = datetime.now()
            
            logger.info(f"Temp file created: {self.temp_path}")
            
        except Exception as e:
            logger.error(f"Error creating temp file: {e}", exc_info=True)
            raise
    
    def _download_google_sheet_as_excel(self):
        """Download Google Sheet as Excel format to temp file."""
        if not self.file_handler or not self.file_handler.sheets_data:
            raise ValueError("No Google Sheets data available from file_handler")
        
        # Create new workbook
        from openpyxl import Workbook
        wb = Workbook()
        wb.remove(wb.active)  # Remove default sheet
        
        # Copy each sheet from Google Sheets data
        for sheet_name, rows in self.file_handler.sheets_data.items():
            ws = wb.create_sheet(title=sheet_name)
            for row_idx, row_data in enumerate(rows, start=1):
                for col_idx, cell_value in enumerate(row_data, start=1):
                    ws.cell(row=row_idx, column=col_idx, value=cell_value)
        
        # Save to temp path
        wb.save(self.temp_path)
        logger.info(f"Google Sheet downloaded as Excel: {self.temp_path}")
    
    def update_inventory(self, item_name: str, value: float, is_addition: bool = False) -> Dict[str, Any]:
        """
        Update inventory for the specified item in temp file.
        
        Args:
            item_name: Name of the inventory item
            value: New value or amount to add/subtract
            is_addition: If True, add/subtract from current value. If False, set to value.
            
        Returns:
            Dictionary with success status and details
        """
        try:
            # Get item location from file_handler
            if not self.file_handler:
                return {
                    "success": False,
                    "message": "No file_handler connected - cannot locate item"
                }
            
            location = self.file_handler.get_item_location(item_name)
            if not location:
                return {
                    "success": False,
                    "message": f"Item not found: {item_name}"
                }
            
            sheet_name = location['sheet']
            row = location['row']
            column = location['column']
            
            logger.debug(f"Updating {item_name} at {sheet_name}[{row},{column}]")
            
            # Get sheet
            if sheet_name not in self.workbook.sheetnames:
                return {
                    "success": False,
                    "message": f"Sheet not found: {sheet_name}"
                }
            
            sheet = self.workbook[sheet_name]
            
            # Get current value
            current_value = sheet.cell(row=row, column=column).value
            if current_value is None:
                current_value = 0
            
            # Calculate new value
            if is_addition:
                new_value = current_value + value
                operation = "add" if value >= 0 else "subtract"
            else:
                new_value = value
                operation = "set"
            
            # Update cell
            sheet.cell(row=row, column=column, value=new_value)
            
            # Save workbook
            self.workbook.save(self.temp_path)
            self.last_save_time = datetime.now()
            
            # Track change
            self.changes_made = True
            change_entry = {
                'timestamp': self.last_save_time.isoformat(),
                'item': item_name,
                'sheet': sheet_name,
                'row': row,
                'column': column,
                'old_value': current_value,
                'new_value': new_value,
                'operation': operation,
                'change_amount': value
            }
            self.change_log.append(change_entry)
            
            logger.info(f"✅ Updated {item_name}: {current_value} → {new_value} (operation: {operation})")
            
            return {
                "success": True,
                "message": f"Updated {item_name}",
                "item": item_name,
                "old_value": current_value,
                "new_value": new_value,
                "operation": operation,
                "temp_file": str(self.temp_path)
            }
            
        except Exception as e:
            error_msg = f"Error updating inventory: {str(e)}"
            logger.error(error_msg, exc_info=True)
            return {
                "success": False,
                "message": error_msg
            }
    
    def batch_update(self, updates: List[Dict[str, Any]]) -> Dict[str, Any]:
        """
        Perform multiple updates in a batch.
        
        Args:
            updates: List of update dicts, each containing:
                - item_name: str
                - value: float
                - is_addition: bool (optional, default False)
                
        Returns:
            Dictionary with batch results
        """
        results = []
        successful = 0
        failed = 0
        
        for update in updates:
            result = self.update_inventory(
                item_name=update['item_name'],
                value=update['value'],
                is_addition=update.get('is_addition', False)
            )
            
            results.append(result)
            if result['success']:
                successful += 1
            else:
                failed += 1
        
        return {
            "success": failed == 0,
            "total": len(updates),
            "successful": successful,
            "failed": failed,
            "results": results
        }
    
    def has_unsaved_changes(self) -> bool:
        """
        Check if there are unsaved changes in temp file.
        
        Returns:
            True if changes have been made since creation
        """
        return self.changes_made
    
    def get_change_log(self) -> List[Dict[str, Any]]:
        """
        Get the log of all changes made.
        
        Returns:
            List of change entries
        """
        return self.change_log.copy()
    
    def get_change_summary(self) -> Dict[str, Any]:
        """
        Get a summary of changes made.
        
        Returns:
            Dictionary with change statistics
        """
        if not self.changes_made:
            return {
                "has_changes": False,
                "total_changes": 0
            }
        
        return {
            "has_changes": True,
            "total_changes": len(self.change_log),
            "first_change": self.change_log[0]['timestamp'] if self.change_log else None,
            "last_change": self.change_log[-1]['timestamp'] if self.change_log else None,
            "last_save": self.last_save_time.isoformat() if self.last_save_time else None,
            "items_modified": len(set(entry['item'] for entry in self.change_log))
        }
    
    def export_change_log(self, output_path: str) -> bool:
        """
        Export change log to JSON file.
        
        Args:
            output_path: Path to save the JSON file
            
        Returns:
            True if successful
        """
        try:
            with open(output_path, 'w') as f:
                json.dump({
                    'source_file': self.source_path,
                    'temp_file': str(self.temp_path),
                    'summary': self.get_change_summary(),
                    'changes': self.change_log
                }, f, indent=2)
            
            logger.info(f"Change log exported to: {output_path}")
            return True
        except Exception as e:
            logger.error(f"Error exporting change log: {e}")
            return False
    
    def get_temp_path(self) -> Path:
        """
        Get the path to the temporary working file.
        
        Returns:
            Path object for temp file
        """
        return self.temp_path
    
    def get_source_path(self) -> str:
        """
        Get the original source file path/URL.
        
        Returns:
            Source path string
        """
        return self.source_path
    
    def discard_changes(self) -> Dict[str, Any]:
        """
        Discard all changes and recreate temp file from source.
        
        Returns:
            Dictionary with success status
        """
        try:
            logger.info("Discarding changes and recreating temp file")
            
            # Close workbook
            if self.workbook:
                self.workbook.close()
            
            # Delete temp file
            if self.temp_path.exists():
                self.temp_path.unlink()
            
            # Reset state
            self.changes_made = False
            self.change_log = []
            
            # Recreate temp file from source
            self._create_temp_file()
            
            return {
                "success": True,
                "message": "Changes discarded, temp file recreated"
            }
            
        except Exception as e:
            error_msg = f"Error discarding changes: {str(e)}"
            logger.error(error_msg, exc_info=True)
            return {
                "success": False,
                "message": error_msg
            }
    
    def close(self):
        """
        Close the temp file and clean up.
        Note: This does NOT save changes back to source!
        Use SyncManager.save_to_source() for that.
        """
        if self.workbook:
            self.workbook.close()
            logger.info("Temp file closed")
    
    def __repr__(self):
        """String representation of TemporaryUpdateHandler."""
        changes = len(self.change_log)
        return f"TemporaryUpdateHandler(source='{Path(self.source_path).name}', temp='{self.temp_path.name}', changes={changes})"
    
    def __del__(self):
        """Cleanup when handler is destroyed."""
        try:
            if self.workbook:
                self.workbook.close()
        except:
            pass
    
    def __enter__(self):
        """Context manager support."""
        return self
    
    def __exit__(self, exc_type, exc_val, exc_tb):
        """Context manager cleanup."""
        self.close()
