"""
NLP search engine module for Dugal Inventory System.
Provides advanced pattern matching and learning capabilities for inventory searches.
"""

import os
import json
import logging
import unicodedata
import re
from typing import Dict, Set, List, Optional, Any
from difflib import SequenceMatcher
from collections import defaultdict
import nltk
from nltk.corpus import wordnet
from nltk.tokenize import word_tokenize
from nltk.stem import WordNetLemmatizer
from openpyxl.utils import column_index_from_string

# Configure logging
logger = logging.getLogger(__name__)

class AdaptiveInventorySearchEngine:
    """Adaptive search engine with learning capabilities."""
    
    def __init__(self, learning_file="dugal_patterns.json", logging_manager=None):
        """Initialize the search engine with learning capabilities."""
        self.learning_file = learning_file
        self.logging_manager = logging_manager
        self.logger = logging.getLogger(__name__)
        self._ensure_nltk_data()
        self.lemmatizer = WordNetLemmatizer()
        self.inventory_cache = {}
        self.synonym_cache = {}
        self.learned_patterns = self._load_learned_patterns()
        self.pending_patterns = defaultdict(int)
        self.confidence_threshold = 3
        self.base_patterns = self._load_base_patterns()
        
        if self.logging_manager:
            self.logging_manager.log_pattern_match({
                'type': 'search_engine_init',
                'learning_file': learning_file,
                'timestamp': datetime.now().isoformat()
            })

        # Auto-registration will be handled by the caller
        # This avoids circular imports with component_manager
        self.logger.debug(f"Search engine initialized (ID: {id(self)})")
        
    def _ensure_nltk_data(self):
        """Ensure required NLTK data is available."""
        try:
            nltk.data.find('tokenizers/punkt')
            nltk.data.find('corpora/wordnet')
        except LookupError:
            nltk.download('punkt', quiet=True)
            nltk.download('wordnet', quiet=True)
            nltk.download('averaged_perceptron_tagger', quiet=True)

    def get_search_engine(self):
        """Get the search engine instance."""
        # Simply return self - component management is handled externally
        # This avoids circular imports with component_manager
        try:
            return self
        except Exception as e:
            self.logger.error(f"Error in get_search_engine: {e}")
            return self  # Always return self for search engine class as last resort

    def _load_base_patterns(self) -> Dict[str, Set[str]]:
        """Load initial known patterns for common inventory items."""
        return {
            # Vodka patterns
            'titos': {'tito', 'titos', "tito's", 'tetos', 'tito s'},
            'ketel': {'kettle', 'ketal', 'kettle one', 'ketel one'},
            'absolut': {'absolute', 'absolut', 'absalut'},
            'stolichnaya': {'stoli', 'stolich', 'stolichnaya'},
            
            # Whiskey patterns
            'buffalo trace': {'buffelo trace', 'buffalo tres', 'buffelo tres'},
            'blanton': {'blantons', "blanton's", 'blantin', 'blanton s'},
            'woodford': {'woodferd', 'woodfort', 'woodford reserve'},
            'maker': {'makers', "maker's", 'maker s'},
            'makers mark': {'makers mark', "maker's mark", 'maker mark'},
            'knob creek': {'nob creek', 'knob creak'},
            
            # Scotch patterns
            'macallan': {'mcallan', 'macallen', 'mcallen'},
            'glenlivet': {'glenlevet', 'glenlivit', 'glen livet'},
            'glenfiddich': {'glenfidich', 'glenfiddic', 'glen fiddich'},
            'laphroaig': {'lafroaig', 'laphroig', 'la froig'},
            
            # Gin patterns
            'hendricks': {'hendrix', 'hendriks', "hendrik's"},
            'tanqueray': {'tanquery', 'tankery', 'tan curry'},
            'bombay': {'bomby', 'bombey', 'bombay sapphire'},
            
            # Tequila patterns
            'patron': {'patrone', 'petronn', 'patron silver'},
            'casamigos': {'casa migos', 'casamigo', 'casa amigos'},
            'don julio': {'donjulio', 'don julios', 'don hulio'},
            
            # Rum patterns
            'bacardi': {'bakardi', 'bacardi', 'backardi'},
            'captain morgan': {'captan morgan', 'captain morgen'},
            
            # Cognac patterns
            'hennessy': {'hennesey', 'henessy', 'henny'},
            'martell': {'martel', 'martell vs', 'martel vs'},
            'remy martin': {'remy', 'remi martin', 'remy martan'},
            
            # Liqueur patterns
            'kahlua': {'kaluha', 'kahlua', 'kaluah'},
            'baileys': {'bailey', 'bailies', "bailey's"},
            'grand marnier': {'grand mariner', 'gran marnier', 'grand marnie'},
            
            # Common descriptors
            'reserve': {'reserva', 'reserved', 'rsv'},
            'special': {'specialty', 'specially', 'spec'},
            'single barrel': {'sig barrel', 'single brl', 'sgl barrel'},
            'limited': {'ltd', 'limit', 'lmtd'},
            'edition': {'ed', 'edt', 'edtn'},
            'straight': {'str', 'strght'},
            'original': {'orig', 'orginal'},
            'signature': {'sig', 'signat'}
        }

    def load_patterns(self, patterns: Dict[str, List[str]]) -> None:
        """Load learned patterns from saved data."""
        logger.debug(f"Loading {len(patterns)} patterns")
        try:
            # Convert patterns to the expected format if needed
            for base_term, variations in patterns.items():
                if isinstance(variations, list):  # Handle list format
                    self.learned_patterns[base_term] = set(variations)
                elif isinstance(variations, set):  # Handle set format
                    self.learned_patterns[base_term] = variations
                else:
                    logger.warning(f"Unexpected pattern format for {base_term}: {type(variations)}")

            if self.logging_manager:
                self.logging_manager.log_pattern_match({
                    'type': 'patterns_loaded',
                    'count': len(patterns),
                    'timestamp': datetime.now().isoformat()
                })

            logger.debug("Patterns loaded successfully")
            
        except Exception as e:
            error_msg = f"Error loading patterns: {e}"
            logger.error(error_msg)
            if self.logging_manager:
                self.logging_manager.log_error(error_msg, {
                    'context': 'pattern_loading',
                    'pattern_count': len(patterns)
                })

    def _load_learned_patterns(self) -> Dict[str, Set[str]]:
        """Load previously learned patterns from file."""
        try:
            if os.path.exists(self.learning_file):
                with open(self.learning_file, 'r') as f:
                    patterns = json.load(f)
                return {k: set(v) for k, v in patterns.items()}
        except Exception as e:
            logger.error(f"Error loading learned patterns: {e}")
        return {}

    def _save_learned_patterns(self) -> None:
        """Save learned patterns to file."""
        try:
            patterns = {k: list(v) for k, v in self.learned_patterns.items()}
            with open(self.learning_file, 'w') as f:
                json.dump(patterns, f, indent=2)
        except Exception as e:
            logger.error(f"Error saving learned patterns: {e}")

    def learn_from_document(self, workbook, search_column: str) -> None:
        """Learn patterns from the current document."""
        logger.debug("Learning patterns from document...")
        if self.logging_manager:
            self.logging_manager.log_pattern_match({
                'type': 'document_learning_start',
                'search_column': search_column,
                'timestamp': datetime.now().isoformat()
            })

        try:
            column_idx = column_index_from_string(search_column)
            new_patterns = defaultdict(set)
            terms_processed = 0
            
            for sheet in workbook.sheetnames:
                ws = workbook[sheet]
                for row in ws.iter_rows(min_row=2, min_col=column_idx, max_col=column_idx):
                    if not row[0].value:
                        continue
                    
                    item_name = str(row[0].value).strip()
                    self._analyze_and_learn_pattern(item_name, new_patterns)
                    terms_processed += 1

            # Add new patterns that appear frequently
            patterns_added = 0
            for base_term, variations in new_patterns.items():
                if base_term not in self.learned_patterns:
                    self.learned_patterns[base_term] = variations
                else:
                    self.learned_patterns[base_term].update(variations)
                patterns_added += 1

            self._save_learned_patterns()
            
            if self.logging_manager:
                self.logging_manager.log_pattern_match({
                    'type': 'document_learning_complete',
                    'terms_processed': terms_processed,
                    'patterns_added': patterns_added,
                    'timestamp': datetime.now().isoformat()
                })
                
            logger.debug(f"Learned {len(new_patterns)} new pattern sets")
            
        except Exception as e:
            error_msg = f"Error learning from document: {e}"
            logger.error(error_msg)
            if self.logging_manager:
                self.logging_manager.log_error(error_msg, {
                    'context': 'document_learning',
                    'search_column': search_column
                })

    def add_term_with_variations(self, base_term: str, variations: List[str]) -> None:
        """Add a new term with its variations."""
        base_norm = self._normalize_text(base_term)
        if base_norm not in self.learned_patterns:
            self.learned_patterns[base_norm] = set()
            
        for variation in variations:
            var_norm = self._normalize_text(variation)
            self.learned_patterns[base_norm].add(var_norm)
            
        self._save_learned_patterns()
        logger.debug(f"Added new term '{base_term}' with {len(variations)} variations")

    def add_manual_pattern(self, base_term: str, variation: str) -> bool:
        """Add a manually defined pattern."""
        try:
            if base_term not in self.learned_patterns:
                self.learned_patterns[base_term] = set()
            self.learned_patterns[base_term].add(variation)
            return True
        except Exception as e:
            logger.error(f"Error adding manual pattern: {e}")
            return False
            
    def learn_pattern(self, pattern: Dict[str, Any]) -> bool:
        """
        Learn a pattern from voice interaction or other components.
        
        Args:
            pattern: A dictionary containing pattern information with at least 'from' and 'to' keys
            
        Returns:
            True if the pattern was successfully learned, False otherwise
        """
        try:
            # Validate pattern
            if not isinstance(pattern, dict) or 'from' not in pattern or 'to' not in pattern:
                logger.warning(f"Invalid pattern format: {pattern}")
                return False
                
            from_term = pattern.get('from', '')
            to_term = pattern.get('to', '')
            confidence = pattern.get('confidence', 0.0)
            
            # Only learn patterns with sufficient confidence
            if confidence < 0.5:
                logger.debug(f"Skipping low confidence pattern: {from_term} -> {to_term} ({confidence})")
                # Add to pending patterns for potential future learning
                self.pending_patterns[f"{from_term}_{to_term}"] += 1
                return False
                
            # Check if this pattern has been seen multiple times in pending patterns
            pending_key = f"{from_term}_{to_term}"
            if pending_key in self.pending_patterns and self.pending_patterns[pending_key] >= 2:
                # Boost confidence for frequently seen patterns
                confidence = max(confidence, 0.7)
                # Clear from pending
                del self.pending_patterns[pending_key]
                
            # Normalize terms
            base_norm = self._normalize_text(to_term)
            var_norm = self._normalize_text(from_term)
            
            # Add the pattern
            if base_norm not in self.learned_patterns:
                self.learned_patterns[base_norm] = set()
                
            self.learned_patterns[base_norm].add(var_norm)
            self._save_learned_patterns()
            
            logger.debug(f"Learned pattern: {from_term} -> {to_term} (confidence: {confidence})")
            
            if self.logging_manager:
                self.logging_manager.log_pattern_match({
                    'type': 'pattern_learned',
                    'from': from_term,
                    'to': to_term,
                    'confidence': confidence,
                    'source': pattern.get('type', 'external'),
                    'timestamp': datetime.now().isoformat()
                })
                
            return True
            
        except Exception as e:
            logger.error(f"Error learning pattern: {e}")
            return False

    def _generate_variations(self, term: str, tokens: List[str]) -> Set[str]:
        """
        Generate variations of a term for improved matching.
        
        Args:
            term: The normalized term to generate variations for
            tokens: The tokenized and processed term
            
        Returns:
            Set[str]: A set of variations of the term
        """
        variations = {term}
        
        try:
            # Add variations from learned patterns
            for token in tokens:
                if token in self.learned_patterns:
                    variations.update(self.learned_patterns[token])
            
            # Add variations from base patterns
            for token in tokens:
                if token in self.base_patterns:
                    variations.update(self.base_patterns[token])
            
            # Add simple plural/singular variants
            if term.endswith('s'):
                variations.add(term[:-1])  # singular form
            else:
                variations.add(term + 's')  # plural form
            
            # Add common misspellings
            for i in range(len(term) - 1):
                # Swap adjacent characters
                swapped = term[:i] + term[i+1] + term[i] + term[i+2:]
                variations.add(swapped)
                
                # Add variants with common character substitutions
                for sub_pair in [('e', 'i'), ('a', 'e'), ('o', 'a'), ('i', 'y')]:
                    if term[i] == sub_pair[0]:
                        subst = term[:i] + sub_pair[1] + term[i+1:]
                        variations.add(subst)
            
            # Handle apostrophes
            if "'" in term:
                variations.add(term.replace("'", ""))
                variations.add(term.replace("'", " "))
            else:
                for i in range(len(term)):
                    if term[i] == 's' and i > 0:
                        variations.add(term[:i] + "'s" + term[i+1:])
            
            return variations
            
        except Exception as e:
            logger.error(f"Error generating variations: {e}")
            # Return at least the original term
            return {term}

    def handle_file_refresh(self, file_path: str) -> None:
        """Handle file refresh events."""
        if self.logging_manager:
            self.logging_manager.log_pattern_match({
                'type': 'file_refresh',
                'file': file_path,
                'timestamp': datetime.now().isoformat()
            })

    def save_current_state(self) -> None:
        """Save current search engine state."""
        if self.logging_manager:
            self.logging_manager.log_pattern_match({
                'type': 'state_save',
                'pattern_count': len(self.learned_patterns),
                'timestamp': datetime.now().isoformat()
            })

    def update_term(self, old_base_term: str, new_base_term: str, new_variations: List[str]) -> None:
        """Update an existing term and its variations."""
        old_norm = self._normalize_text(old_base_term)
        new_norm = self._normalize_text(new_base_term)
        
        # Remove old term
        if old_norm in self.learned_patterns:
            del self.learned_patterns[old_norm]
            
        # Add new term and variations
        self.learned_patterns[new_norm] = set(self._normalize_text(v) for v in new_variations)
        self._save_learned_patterns()
        logger.debug(f"Updated term '{old_base_term}' to '{new_base_term}'")

    def build_search_index(self, workbook, search_column: str) -> None:
        """Build search index from workbook."""
        logger.debug("Building search index...")
        self.inventory_cache.clear()
        self.workbook = workbook  # Store reference to workbook
        column_idx = column_index_from_string(search_column)
        
        # Find input column index by looking for "Ending Inventory: FARMbar" in row 1
        if not hasattr(self, 'input_column_index') or not self.input_column_index:
            self.input_column_index = self._find_inventory_column(workbook)
            if self.input_column_index:
                logger.debug(f"Found input column 'Ending Inventory: FARMbar' at index {self.input_column_index}")
            else:
                # Only use default if we couldn't find the column
                self.input_column_index = 6  # Default fallback
                logger.warning(f"Could not find 'Ending Inventory: FARMbar' column, using default column {self.input_column_index}")
        
        for sheet in workbook.sheetnames:
            ws = workbook[sheet]
            self._index_sheet(ws, sheet, column_idx)
        
        logger.debug(f"Indexed {len(self.inventory_cache)} unique terms")
        
        # Register in global registry after indexing
        try:
            from global_registry import GlobalRegistry
            GlobalRegistry.register('search_engine', self)
            logger.debug("Registered indexed search engine in global registry")
        except Exception as e:
            logger.error(f"Error registering search engine: {e}")

    def _find_inventory_column(self, workbook) -> int:
        """Find the column index for 'Ending Inventory: FARMbar'."""
        try:
            # Try to find the column in the first sheet's header row
            first_sheet = workbook[workbook.sheetnames[0]]
            for cell in first_sheet[1]:  # Row 1 (header row)
                if cell.value and "Ending Inventory: FARMbar" in str(cell.value):
                    self.input_column_name = str(cell.value)
                    return cell.column
            
            # If not found in first sheet, try others
            for sheet_name in workbook.sheetnames:
                sheet = workbook[sheet_name]
                for cell in sheet[1]:
                    if cell.value and "Ending Inventory: FARMbar" in str(cell.value):
                        self.input_column_name = str(cell.value)
                        return cell.column
                        
            # If still not found, return None
            return None
        except Exception as e:
            logger.error(f"Error finding inventory column: {e}")
            return None

    def _index_sheet(self, sheet, sheet_name: str, column_idx: int) -> None:
        """Index a single sheet."""
        for row in sheet.iter_rows(min_row=2, min_col=column_idx, max_col=column_idx):
            if not row[0].value:
                continue
                
            item_name = str(row[0].value).strip()
            normalized_name = self._normalize_text(item_name)
            tokens = self._tokenize_and_process(normalized_name)
            
            cache_entry = {
                'original': item_name,
                'sheet': sheet_name,
                'row': row[0].row,
                'tokens': tokens,
                'variations': self._generate_variations(normalized_name, tokens)
            }
            
            self._add_to_cache(normalized_name, cache_entry)

    def _add_to_cache(self, key: str, entry: Dict[str, Any]) -> None:
        """Add an entry to the search cache."""
        if key not in self.inventory_cache:
            self.inventory_cache[key] = []
        self.inventory_cache[key].append(entry)
        
        # Add variations to cache
        for variant in entry['variations']:
            if variant not in self.inventory_cache:
                self.inventory_cache[variant] = []
            self.inventory_cache[variant].append(entry)

    def register_in_registry(self):
        """Register this search engine instance in the global registry."""
        try:
            from global_registry import GlobalRegistry
            GlobalRegistry.register('search_engine', self)
            self.logger.debug(f"Search engine registered in global registry (ID: {id(self)})")
            return True
        except Exception as e:
            self.logger.error(f"Error registering in global registry: {e}")
            return False

    def connect_to_workbook(self, workbook, search_column=None):
        """Connect search engine to a workbook and optionally index it."""
        try:
            self.workbook = workbook
            self.logger.info(f"Connected to workbook: {id(workbook)}")
            
            if search_column:
                self.search_column = search_column
                self.build_search_index(workbook, search_column)
                
            # Register in global registry after updating
            self.register_in_registry()
            
            return True
        except Exception as e:
            self.logger.error(f"Error connecting to workbook: {e}")
            return False

    def search(self, sheet, search_term: str, threshold: float = 0.4) -> Optional[int]:
        """
        Search for an item in the specified sheet.
        
        Args:
            sheet: Worksheet to search in
            search_term: Term to search for
            threshold: Minimum similarity threshold
            
        Returns:
            Optional[int]: Matching row number or None
        """
        normalized_search = self._normalize_text(search_term)
        matches = self._search_with_variations(normalized_search, threshold)
        
        # Filter matches for current sheet
        sheet_matches = [m for m in matches if m['sheet'] == sheet.title]
        if sheet_matches:
            return sheet_matches[0]['row']
            
        return None

    def _search_with_variations(self, search_term: str, threshold: float) -> List[Dict]:
        """Search using all possible variations of the term with enhanced matching."""
        matches = []
        seen_items = set()
        
        # Log the search attempt
        self.logger.debug(f"Fuzzy searching for '{search_term}' with threshold {threshold}")
        
        # Get variations from learned patterns
        search_variations = {search_term}
        tokens = self._tokenize_and_process(search_term)
        
        # Add known variations
        for token in tokens:
            if token in self.learned_patterns:
                search_variations.update(self.learned_patterns[token])
        
        # Add partial search variations - focus on first words
        search_tokens = search_term.split()
        if len(search_tokens) > 1:
            search_variations.add(search_tokens[0])  # First word only
            if len(search_tokens) > 2:
                search_variations.add(f"{search_tokens[0]} {search_tokens[1]}")  # First two words
        
        # Search using all variations
        for variation in search_variations:
            self.logger.debug(f"Trying variation: '{variation}'")
            
            for cached_item, entries in self.inventory_cache.items():
                for entry in entries:
                    if entry['original'] in seen_items:
                        continue
                        
                    similarity = self._calculate_enhanced_similarity(
                        variation,
                        cached_item,
                        entry['tokens'] if 'tokens' in entry else None
                    )
                    
                    if similarity >= threshold:
                        matches.append({
                            'item': entry['original'],
                            'sheet': entry['sheet'],
                            'row': entry['row'],
                            'confidence': similarity,
                            'matched_term': variation
                        })
                        seen_items.add(entry['original'])
        
        # Sort by confidence
        matches.sort(key=lambda x: x['confidence'], reverse=True)
        
        # Log match results
        if matches:
            self.logger.debug(f"Found {len(matches)} fuzzy matches. Best match: '{matches[0]['item']}' with score {matches[0]['confidence']:.2f} from '{matches[0]['matched_term']}'")
        else:
            self.logger.debug(f"No fuzzy matches found for '{search_term}'")
        
        return matches

    def _calculate_enhanced_similarity(self, term1, term2, tokens=None):
        """Calculate similarity with improved partial matching."""
        # Use existing similarity calculation if available
        if hasattr(self, '_calculate_similarity') and callable(getattr(self, '_calculate_similarity')):
            base_similarity = self._calculate_similarity(term1, term2, tokens)
            if base_similarity >= 0.8:  # If it's already a good match
                return base_similarity
        else:
            base_similarity = 0
        
        # Simple exact match
        if term1 == term2:
            return 1.0
            
        # Check if one is a substring of the other
        if term1 in term2:
            return max(base_similarity, 0.9 * (len(term1) / len(term2)))
        if term2 in term1:
            return max(base_similarity, 0.9 * (len(term2) / len(term1)))
        
        # Token-based matching for multi-word terms
        term1_tokens = term1.split()
        term2_tokens = term2.split()
        
        # Check for token matches
        matching_tokens = sum(1 for t in term1_tokens if any(t in t2 or t2 in t for t2 in term2_tokens))
        if matching_tokens > 0:
            token_similarity = 0.7 * (matching_tokens / max(len(term1_tokens), len(term2_tokens)))
            return max(base_similarity, token_similarity)
        
        # If tokens provided and original calculation was low, try with Levenshtein
        return max(base_similarity, self._levenshtein_similarity(term1, term2))

    def _levenshtein_similarity(self, s1, s2):
        """Calculate normalized Levenshtein similarity."""
        if not s1 or not s2:
            return 0.0
            
        # Calculate Levenshtein distance
        m, n = len(s1), len(s2)
        d = [[0 for _ in range(n+1)] for _ in range(m+1)]
        
        for i in range(m+1):
            d[i][0] = i
        for j in range(n+1):
            d[0][j] = j
            
        for j in range(1, n+1):
            for i in range(1, m+1):
                cost = 0 if s1[i-1] == s2[j-1] else 1
                d[i][j] = min(
                    d[i-1][j] + 1,      # deletion
                    d[i][j-1] + 1,      # insertion
                    d[i-1][j-1] + cost  # substitution
                )
        
        distance = d[m][n]
        max_len = max(m, n)
        if max_len == 0:
            return 1.0
        
        return 1.0 - (distance / max_len)

    def find_item(self, search_term: str) -> Dict[str, Any]:
        """Find an item in the inventory and return its information."""
        self.logger = logging.getLogger(__name__)

        try:
            # Normalize search term
            search_term = search_term.lower().strip()
            
            self.logger.debug(f"Searching for inventory item: {search_term}")
            
            # Log cache state for debugging
            if hasattr(self, 'inventory_cache'):
                self.logger.debug(f"Inventory cache contains {len(self.inventory_cache)} items")
                sample_keys = list(self.inventory_cache.keys())[:5] if self.inventory_cache else []
                self.logger.debug(f"Sample cache keys: {sample_keys}")
            else:
                self.logger.debug("No inventory_cache attribute found")
            
            # Try to find the item in the index
            if not hasattr(self, 'inventory_cache') or not self.inventory_cache:
                self.logger.warning("Search index not initialized")
                return {'found': False, 'message': 'Search index not initialized'}
            
            # Check for direct match
            if search_term in self.inventory_cache:
                entries = self.inventory_cache[search_term]
                if entries:
                    first_entry = entries[0]
                    result = {
                        'found': True,
                        'item': first_entry['original'],
                        'sheet': first_entry['sheet'],
                        'row': first_entry['row'],
                        'value': self._get_value_for_item(first_entry)
                    }
                    self.logger.debug(f"Direct match found: {result}")
                    return result
            
            # If not found directly, log this fact
            self.logger.debug(f"No direct match for '{search_term}', trying similar items")
            
            # Check for similar items
            similar_items = self._search_with_variations(search_term, 0.60)
            if similar_items and len(similar_items) > 0:
                best_match = similar_items[0]
                result = {
                    'found': True,
                    'item': best_match['item'],
                    'sheet': best_match['sheet'],
                    'row': best_match['row'],
                    'value': self._get_value_for_item(best_match),
                    'was_similar': True
                }
                self.logger.debug(f"Similar match found: {result}")
                return result
            
            self.logger.debug(f"No matches found for '{search_term}'")
            return {'found': False, 'message': 'Item not found in inventory'}
        
        except Exception as e:
            self.logger.error(f"Error finding item: {e}")
            self.logger.exception("Detailed search error:")  # This logs the full stack trace
            return {'found': False, 'message': f"Error finding item: {str(e)}"}
        
    def _get_value_for_item(self, item_info):
        """Get the current value for an item from the workbook."""
        try:
            if not hasattr(self, 'workbook') or not self.workbook:
                return 'unknown'
            
            sheet_name = item_info.get('sheet')
            row = item_info.get('row')
            
            if not sheet_name or not row:
                return 'unknown'
            
            # Get the sheet
            sheet = self.workbook[sheet_name]
            
            # Determine the value column (this depends on your structure)
            # Typically this would be determined by your input_column setting
            value_col = self._get_value_column()
            
            if not value_col:
                return 'unknown'
            
            # Get the value
            value = sheet.cell(row=row, column=value_col).value
            
            # Convert to float if possible
            try:
                return float(value) if value is not None else 0
            except (ValueError, TypeError):
                return str(value) if value is not None else 'unknown'
                
        except Exception as e:
            logger.error(f"Error getting value for item: {e}")
            return 'unknown'
        
    def _get_value_column(self):
        """Get the column index for the value column."""
        # This should be based on your application's configuration
        # For example, getting it from the excel_handler or a config setting
        if hasattr(self, 'input_column_index'):
            return self.input_column_index
        return None

    def _get_current_value(self, item_data: Dict[str, Any]) -> float:
        """Get the current value of an item from its data."""
        try:
            if 'values' in item_data and item_data['values']:
                return item_data['values'][0]  # Return the first (most recent) value
            return 'unknown'
        except:
            return 'unknown'

    def _normalize_text(self, text: str) -> str:
        """Normalize text for consistent comparison."""
        text = unicodedata.normalize('NFKD', text.lower().strip())
        text = re.sub(r'[^\w\s-]', '', text)
        return text

    def _tokenize_and_process(self, text: str) -> List[str]:
        """Tokenize and process text."""
        tokens = word_tokenize(text)
        processed = []
        
        for token in tokens:
            lemma = self.lemmatizer.lemmatize(token)
            processed.append(lemma)
            
            # Add WordNet synonyms
            if token not in self.synonym_cache:
                synsets = wordnet.synsets(token)
                synonyms = {lemma.name() for synset in synsets for lemma in synset.lemmas()}
                self.synonym_cache[token] = synonyms
            
            processed.extend(self.synonym_cache[token])
            
        return processed

    def _calculate_similarity(self, term1: str, term2: str, term2_tokens: List[str]) -> float:
        """Calculate similarity between terms."""
        # Direct match
        if term1 == term2:
            return 1.0
            
        # Base similarity
        sequence_ratio = SequenceMatcher(None, term1, term2).ratio()
        
        # Token similarity
        term1_tokens = self._tokenize_and_process(term1)
        token_overlap = len(set(term1_tokens) & set(term2_tokens)) / max(len(term1_tokens), len(term2_tokens))
        
        # Pattern matching
        pattern_match = 0.0
        if term1 in self.learned_patterns and term2 in self.learned_patterns[term1]:
            pattern_match = 0.2
            
        # Combined score
        return (sequence_ratio * 0.4) + (token_overlap * 0.4) + pattern_match

    def diagnose_search_index(self):
        """Diagnose the search index state and report issues."""
        logger.debug("=== SEARCH ENGINE DIAGNOSTIC ===")
        logger.debug(f"Search engine instance ID: {id(self)}")
        
        # Check workbook reference
        if hasattr(self, 'workbook') and self.workbook:
            logger.debug(f"Workbook reference exists: {bool(self.workbook)}")
            if hasattr(self.workbook, 'sheetnames'):
                logger.debug(f"Workbook contains {len(self.workbook.sheetnames)} sheets")
        else:
            logger.error("No workbook reference in search engine")
        
        # Check inventory cache
        if hasattr(self, 'inventory_cache'):
            cache_size = len(self.inventory_cache)
            logger.debug(f"Inventory cache contains {cache_size} items")
            if cache_size > 0:
                logger.debug(f"Sample keys: {list(self.inventory_cache.keys())[:3]}")
        else:
            logger.error("No inventory cache in search engine")
        
        # Check input column
        if hasattr(self, 'input_column_index'):
            logger.debug(f"Input column index: {self.input_column_index}")
        else:
            logger.warning("No input column index in search engine")
        
        if hasattr(self, 'input_column_name'):
            logger.debug(f"Input column name: {self.input_column_name}")
        else:
            logger.warning("No input column name in search engine")
        
        # Check search column
        if hasattr(self, 'search_column'):
            logger.debug(f"Search column: {self.search_column}")
        else:
            logger.warning("No search column in search engine")
        
        logger.debug("=== END DIAGNOSTIC ===")
        return True
            
    def get_all_patterns(self) -> Dict[str, List[str]]:
        """Get all learned patterns."""
        return {
            term: list(variations) 
            for term, variations in self.learned_patterns.items()
        }
