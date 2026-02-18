"""
UniversalFileHandler - Multi-Format File Operations
Handles opening, reading, and searching files from multiple sources:
- Excel files (.xlsx, .xlsm) from local/OneDrive
- Google Sheets via Google Sheets API
- CSV files
NO WRITE OPERATIONS - All writes go through TemporaryUpdateHandler.
"""

import logging
import csv
from pathlib import Path
from typing import Dict, Any, List, Optional, Union
from openpyxl import load_workbook
from openpyxl.workbook import Workbook
import re

logger = logging.getLogger(__name__)


class UniversalFileHandler:
    """
    Universal file handler supporting Excel, Google Sheets, and CSV.
    
    Responsibilities:
    - Open files from multiple sources (read-only)
    - Normalize data format across file types
    - Provide file metadata
    - Interface with search engine
    - NO WRITE OPERATIONS
    """
    
    # File type constants
    EXCEL = "excel"
    GOOGLE_SHEETS = "google_sheets"
    CSV = "csv"
    UNKNOWN = "unknown"
    
    def __init__(self, search_engine=None, google_credentials_path=None):
        """
        Initialize UniversalFileHandler.
        
        Args:
            search_engine: Optional search engine instance for inventory lookups
            google_credentials_path: Path to Google service account JSON (optional)
        """
        self.file_path = None
        self.file_type = None
        self.workbook = None  # For Excel files
        self.sheets_data = None  # For Google Sheets
        self.csv_data = None  # For CSV files
        self.search_engine = search_engine
        self.sheet_names = []
        self.file_info = {}
        
        # Google Sheets setup
        self.google_credentials_path = google_credentials_path
        self.google_service = None
        self._init_google_sheets()
        
        logger.debug("UniversalFileHandler initialized (read-only mode)")
    
    def _init_google_sheets(self):
        """Initialize Google Sheets API if credentials are available."""
        if not self.google_credentials_path:
            logger.debug("No Google credentials provided - Google Sheets support disabled")
            return
        
        try:
            from google.oauth2 import service_account
            from googleapiclient.discovery import build
            
            creds = service_account.Credentials.from_service_account_file(
                self.google_credentials_path,
                scopes=['https://www.googleapis.com/auth/spreadsheets.readonly']
            )
            self.google_service = build('sheets', 'v4', credentials=creds)
            logger.info("Google Sheets API initialized successfully")
        except ImportError:
            logger.warning("Google API libraries not installed. Install with: pip install google-auth google-api-python-client")
        except Exception as e:
            logger.error(f"Failed to initialize Google Sheets API: {e}")
    
    def open_file(self, file_path_or_url: str) -> Dict[str, Any]:
        """
        Smart open - automatically detects file type and opens appropriately.
        
        Args:
            file_path_or_url: File path, Google Sheets URL, or spreadsheet ID
            
        Returns:
            Dictionary with success status and file info
        """
        # Detect file type
        file_type = self._detect_file_type(file_path_or_url)
        
        if file_type == self.EXCEL:
            return self.open_excel(file_path_or_url)
        elif file_type == self.GOOGLE_SHEETS:
            return self.open_google_sheet(file_path_or_url)
        elif file_type == self.CSV:
            return self.open_csv(file_path_or_url)
        else:
            return {
                "success": False,
                "message": f"Unknown file type: {file_path_or_url}"
            }
    
    def _detect_file_type(self, file_path_or_url: str) -> str:
        """
        Detect the type of file/URL.
        
        Args:
            file_path_or_url: Path or URL to analyze
            
        Returns:
            File type constant (EXCEL, GOOGLE_SHEETS, CSV, UNKNOWN)
        """
        # Google Sheets URL patterns
        if "docs.google.com/spreadsheets" in file_path_or_url:
            return self.GOOGLE_SHEETS
        
        # If it looks like a spreadsheet ID (alphanumeric, ~44 chars)
        if re.match(r'^[a-zA-Z0-9_-]{40,50}$', file_path_or_url):
            return self.GOOGLE_SHEETS
        
        # File extension detection
        path = Path(file_path_or_url)
        suffix = path.suffix.lower()
        
        if suffix in ['.xlsx', '.xlsm']:
            return self.EXCEL
        elif suffix == '.csv':
            return self.CSV
        
        return self.UNKNOWN
    
    def open_excel(self, file_path: str) -> Dict[str, Any]:
        """
        Open an Excel file in read-only mode.
        
        Args:
            file_path: Full path to the Excel file
            
        Returns:
            Dictionary with success status and file info
        """
        try:
            # Validate path
            path = Path(file_path)
            if not path.exists():
                return {
                    "success": False,
                    "message": f"File not found: {file_path}"
                }
            
            if not path.suffix.lower() in ['.xlsx', '.xlsm']:
                return {
                    "success": False,
                    "message": f"Invalid Excel file type: {path.suffix}"
                }
            
            # Store info
            self.file_path = str(path)
            self.file_type = self.EXCEL
            
            # Open workbook READ-ONLY
            logger.info(f"Opening Excel file (read-only): {self.file_path}")
            self.workbook = load_workbook(self.file_path, read_only=True, data_only=True)
            
            # Extract file info
            self.sheet_names = self.workbook.sheetnames
            self.file_info = {
                'path': self.file_path,
                'filename': path.name,
                'type': self.EXCEL,
                'size_bytes': path.stat().st_size,
                'sheet_count': len(self.sheet_names),
                'sheets': self.sheet_names
            }
            
            logger.info(f"Excel file opened successfully: {len(self.sheet_names)} sheets found")
            
            return {
                "success": True,
                "message": "Excel file opened successfully",
                "file_info": self.file_info
            }
            
        except PermissionError:
            error_msg = f"Permission denied: {file_path}"
            logger.error(error_msg)
            return {"success": False, "message": error_msg}
        
        except Exception as e:
            error_msg = f"Error opening Excel file: {str(e)}"
            logger.error(error_msg, exc_info=True)
            return {"success": False, "message": error_msg}
    
    def open_google_sheet(self, spreadsheet_id_or_url: str) -> Dict[str, Any]:
        """
        Open a Google Sheet in read-only mode.
        
        Args:
            spreadsheet_id_or_url: Spreadsheet ID or full Google Sheets URL
            
        Returns:
            Dictionary with success status and file info
        """
        if not self.google_service:
            return {
                "success": False,
                "message": "Google Sheets API not initialized. Provide credentials_path when creating handler."
            }
        
        try:
            # Extract spreadsheet ID from URL if needed
            spreadsheet_id = self._extract_spreadsheet_id(spreadsheet_id_or_url)
            
            if not spreadsheet_id:
                return {
                    "success": False,
                    "message": "Invalid Google Sheets URL or ID"
                }
            
            logger.info(f"Opening Google Sheet (read-only): {spreadsheet_id}")
            
            # Get spreadsheet metadata
            spreadsheet = self.google_service.spreadsheets().get(
                spreadsheetId=spreadsheet_id
            ).execute()
            
            # Extract sheet names
            self.sheet_names = [sheet['properties']['title'] for sheet in spreadsheet['sheets']]
            
            # Load all sheet data
            self.sheets_data = {}
            for sheet_name in self.sheet_names:
                result = self.google_service.spreadsheets().values().get(
                    spreadsheetId=spreadsheet_id,
                    range=sheet_name
                ).execute()
                self.sheets_data[sheet_name] = result.get('values', [])
            
            # Store info
            self.file_path = f"https://docs.google.com/spreadsheets/d/{spreadsheet_id}/edit"
            self.file_type = self.GOOGLE_SHEETS
            self.file_info = {
                'path': self.file_path,
                'filename': spreadsheet.get('properties', {}).get('title', 'Untitled'),
                'type': self.GOOGLE_SHEETS,
                'spreadsheet_id': spreadsheet_id,
                'sheet_count': len(self.sheet_names),
                'sheets': self.sheet_names
            }
            
            logger.info(f"Google Sheet opened successfully: {len(self.sheet_names)} sheets found")
            
            return {
                "success": True,
                "message": "Google Sheet opened successfully",
                "file_info": self.file_info
            }
            
        except Exception as e:
            error_msg = f"Error opening Google Sheet: {str(e)}"
            logger.error(error_msg, exc_info=True)
            return {"success": False, "message": error_msg}
    
    def _extract_spreadsheet_id(self, url_or_id: str) -> Optional[str]:
        """
        Extract spreadsheet ID from URL or return ID if already provided.
        
        Args:
            url_or_id: Google Sheets URL or spreadsheet ID
            
        Returns:
            Spreadsheet ID or None if invalid
        """
        # If it's already an ID (alphanumeric string)
        if re.match(r'^[a-zA-Z0-9_-]{40,50}$', url_or_id):
            return url_or_id
        
        # Extract from URL
        match = re.search(r'/spreadsheets/d/([a-zA-Z0-9_-]+)', url_or_id)
        if match:
            return match.group(1)
        
        return None
    
    def open_csv(self, file_path: str) -> Dict[str, Any]:
        """
        Open a CSV file in read-only mode.
        
        Args:
            file_path: Full path to the CSV file
            
        Returns:
            Dictionary with success status and file info
        """
        try:
            # Validate path
            path = Path(file_path)
            if not path.exists():
                return {
                    "success": False,
                    "message": f"File not found: {file_path}"
                }
            
            if not path.suffix.lower() == '.csv':
                return {
                    "success": False,
                    "message": f"Invalid CSV file type: {path.suffix}"
                }
            
            logger.info(f"Opening CSV file (read-only): {file_path}")
            
            # Read CSV data
            with open(file_path, 'r', encoding='utf-8') as f:
                reader = csv.reader(f)
                self.csv_data = list(reader)
            
            # Store info
            self.file_path = str(path)
            self.file_type = self.CSV
            self.sheet_names = ['Sheet1']  # CSV has one implicit sheet
            self.file_info = {
                'path': self.file_path,
                'filename': path.name,
                'type': self.CSV,
                'size_bytes': path.stat().st_size,
                'sheet_count': 1,
                'sheets': self.sheet_names,
                'row_count': len(self.csv_data)
            }
            
            logger.info(f"CSV file opened successfully: {len(self.csv_data)} rows found")
            
            return {
                "success": True,
                "message": "CSV file opened successfully",
                "file_info": self.file_info
            }
            
        except Exception as e:
            error_msg = f"Error opening CSV file: {str(e)}"
            logger.error(error_msg, exc_info=True)
            return {"success": False, "message": error_msg}
    
    def get_file_info(self) -> Dict[str, Any]:
        """
        Get information about the currently open file.
        
        Returns:
            Dictionary with file metadata
        """
        if not self.file_type:
            return {
                "success": False,
                "message": "No file is currently open"
            }
        
        return {
            "success": True,
            "file_info": self.file_info
        }
    
    def get_sheet_names(self) -> List[str]:
        """
        Get list of sheet names in the file.
        
        Returns:
            List of sheet names, or empty list if no file open
        """
        return self.sheet_names if self.file_type else []
    
    def connect_search_engine(self, search_engine):
        """
        Connect a search engine for inventory lookups.
        
        Args:
            search_engine: Search engine instance with find_item() method
        """
        self.search_engine = search_engine
        logger.debug("Search engine connected to UniversalFileHandler")
    
    def search_item(self, item_name: str) -> Dict[str, Any]:
        """
        Search for an inventory item (read-only).
        
        Args:
            item_name: Name of the item to search for
            
        Returns:
            Dictionary with search results
        """
        if not self.search_engine:
            return {
                "success": False,
                "found": False,
                "message": "No search engine connected"
            }
        
        try:
            result = self.search_engine.find_item(item_name)
            return {
                "success": True,
                **result
            }
        except Exception as e:
            logger.error(f"Error searching for item '{item_name}': {e}")
            return {
                "success": False,
                "found": False,
                "message": str(e)
            }
    
    def get_item_value(self, item_name: str) -> Optional[float]:
        """
        Get the current value of an inventory item (read-only).
        
        Args:
            item_name: Name of the item
            
        Returns:
            Current value, or None if not found
        """
        result = self.search_item(item_name)
        if result.get('found'):
            return result.get('value', 0)
        return None
    
    def get_item_location(self, item_name: str) -> Optional[Dict[str, Any]]:
        """
        Get the location of an item in the file (sheet, row, column).
        
        Args:
            item_name: Name of the item
            
        Returns:
            Dictionary with location info, or None if not found
        """
        result = self.search_item(item_name)
        if result.get('found'):
            return {
                'sheet': result.get('sheet'),
                'row': result.get('row'),
                'column': result.get('column'),
                'item': result.get('item')
            }
        return None
    
    def close_file(self):
        """
        Close the currently open file.
        """
        if self.file_type == self.EXCEL and self.workbook:
            self.workbook.close()
        
        logger.info(f"Closed file: {self.file_path}")
        
        # Reset state
        self.workbook = None
        self.sheets_data = None
        self.csv_data = None
        self.file_path = None
        self.file_type = None
        self.sheet_names = []
        self.file_info = {}
    
    def is_file_open(self) -> bool:
        """
        Check if a file is currently open.
        
        Returns:
            True if file is open, False otherwise
        """
        return self.file_type is not None
    
    def get_file_path(self) -> Optional[str]:
        """
        Get the path/URL of the currently open file.
        
        Returns:
            File path or URL, or None if no file open
        """
        return self.file_path
    
    def get_file_type(self) -> Optional[str]:
        """
        Get the type of the currently open file.
        
        Returns:
            File type constant (EXCEL, GOOGLE_SHEETS, CSV), or None
        """
        return self.file_type
    
    def reload_file(self) -> Dict[str, Any]:
        """
        Reload the file from its source (useful if external changes were made).
        
        Returns:
            Dictionary with success status
        """
        if not self.file_path:
            return {
                "success": False,
                "message": "No file is currently open"
            }
        
        # Store path and type
        path = self.file_path
        file_type = self.file_type
        
        # Close current file
        self.close_file()
        
        # Reopen based on type
        if file_type == self.EXCEL:
            return self.open_excel(path)
        elif file_type == self.GOOGLE_SHEETS:
            return self.open_google_sheet(path)
        elif file_type == self.CSV:
            return self.open_csv(path)
        
        return {
            "success": False,
            "message": f"Unknown file type: {file_type}"
        }
    
    def __repr__(self):
        """String representation of UniversalFileHandler."""
        if self.file_type:
            filename = Path(self.file_path).name if self.file_type != self.GOOGLE_SHEETS else self.file_info.get('filename', 'Unknown')
            return f"UniversalFileHandler(file='{filename}', type={self.file_type}, sheets={len(self.sheet_names)}, read_only=True)"
        return "UniversalFileHandler(no file open)"
    
    def __del__(self):
        """Cleanup when UniversalFileHandler is destroyed."""
        if self.file_type == self.EXCEL and self.workbook:
            try:
                self.workbook.close()
            except:
                pass
