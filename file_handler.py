"""
FileHandler - Read-Only File Operations
Handles opening, reading, and searching Excel files from OneDrive.
NO WRITE OPERATIONS - All writes go through TemporaryUpdateHandler.
"""

import logging
from pathlib import Path
from typing import Dict, Any, List, Optional
from openpyxl import load_workbook
from openpyxl.workbook import Workbook

logger = logging.getLogger(__name__)


class FileHandler:
    """
    Read-only handler for Excel files from OneDrive.
    
    Responsibilities:
    - Open files from OneDrive (read-only)
    - Provide file metadata
    - Interface with search engine
    - NO WRITE OPERATIONS
    """
    
    def __init__(self, search_engine=None):
        """
        Initialize FileHandler.
        
        Args:
            search_engine: Optional search engine instance for inventory lookups
        """
        self.onedrive_path = None
        self.workbook = None
        self.search_engine = search_engine
        self.sheet_names = []
        self.file_info = {}
        
        logger.debug("FileHandler initialized (read-only mode)")
    
    def open_file(self, onedrive_path: str) -> Dict[str, Any]:
        """
        Open an Excel file from OneDrive in read-only mode.
        
        Args:
            onedrive_path: Full path to the file on OneDrive
            
        Returns:
            Dictionary with success status and file info
        """
        try:
            # Validate path
            file_path = Path(onedrive_path)
            if not file_path.exists():
                return {
                    "success": False,
                    "message": f"File not found: {onedrive_path}"
                }
            
            if not file_path.suffix.lower() in ['.xlsx', '.xlsm']:
                return {
                    "success": False,
                    "message": f"Invalid file type: {file_path.suffix}"
                }
            
            # Store path
            self.onedrive_path = str(file_path)
            
            # Open workbook READ-ONLY
            logger.info(f"Opening file (read-only): {self.onedrive_path}")
            self.workbook = load_workbook(self.onedrive_path, read_only=True, data_only=True)
            
            # Extract file info
            self.sheet_names = self.workbook.sheetnames
            self.file_info = {
                'path': self.onedrive_path,
                'filename': file_path.name,
                'size_bytes': file_path.stat().st_size,
                'sheet_count': len(self.sheet_names),
                'sheets': self.sheet_names
            }
            
            logger.info(f"File opened successfully: {len(self.sheet_names)} sheets found")
            
            return {
                "success": True,
                "message": "File opened successfully",
                "file_info": self.file_info
            }
            
        except PermissionError:
            error_msg = f"Permission denied: {onedrive_path}"
            logger.error(error_msg)
            return {"success": False, "message": error_msg}
        
        except Exception as e:
            error_msg = f"Error opening file: {str(e)}"
            logger.error(error_msg, exc_info=True)
            return {"success": False, "message": error_msg}
    
    def get_file_info(self) -> Dict[str, Any]:
        """
        Get information about the currently open file.
        
        Returns:
            Dictionary with file metadata
        """
        if not self.workbook:
            return {
                "success": False,
                "message": "No file is currently open"
            }
        
        return {
            "success": True,
            "file_info": self.file_info
        }
    
    def get_sheet_names(self) -> List[str]:
        """
        Get list of sheet names in the workbook.
        
        Returns:
            List of sheet names, or empty list if no file open
        """
        return self.sheet_names if self.workbook else []
    
    def connect_search_engine(self, search_engine):
        """
        Connect a search engine for inventory lookups.
        
        Args:
            search_engine: Search engine instance with find_item() method
        """
        self.search_engine = search_engine
        logger.debug("Search engine connected to FileHandler")
    
    def search_item(self, item_name: str) -> Dict[str, Any]:
        """
        Search for an inventory item (read-only).
        
        Args:
            item_name: Name of the item to search for
            
        Returns:
            Dictionary with search results
        """
        if not self.search_engine:
            return {
                "success": False,
                "found": False,
                "message": "No search engine connected"
            }
        
        try:
            result = self.search_engine.find_item(item_name)
            return {
                "success": True,
                **result
            }
        except Exception as e:
            logger.error(f"Error searching for item '{item_name}': {e}")
            return {
                "success": False,
                "found": False,
                "message": str(e)
            }
    
    def get_item_value(self, item_name: str) -> Optional[float]:
        """
        Get the current value of an inventory item (read-only).
        
        Args:
            item_name: Name of the item
            
        Returns:
            Current value, or None if not found
        """
        result = self.search_item(item_name)
        if result.get('found'):
            return result.get('value', 0)
        return None
    
    def get_item_location(self, item_name: str) -> Optional[Dict[str, Any]]:
        """
        Get the location of an item in the workbook (sheet, row, column).
        
        Args:
            item_name: Name of the item
            
        Returns:
            Dictionary with location info, or None if not found
        """
        result = self.search_item(item_name)
        if result.get('found'):
            return {
                'sheet': result.get('sheet'),
                'row': result.get('row'),
                'column': result.get('column'),
                'item': result.get('item')
            }
        return None
    
    def close_file(self):
        """
        Close the currently open file.
        """
        if self.workbook:
            self.workbook.close()
            logger.info(f"Closed file: {self.onedrive_path}")
            self.workbook = None
            self.onedrive_path = None
            self.sheet_names = []
            self.file_info = {}
    
    def is_file_open(self) -> bool:
        """
        Check if a file is currently open.
        
        Returns:
            True if file is open, False otherwise
        """
        return self.workbook is not None
    
    def get_onedrive_path(self) -> Optional[str]:
        """
        Get the OneDrive path of the currently open file.
        
        Returns:
            OneDrive path, or None if no file open
        """
        return self.onedrive_path
    
    def reload_file(self) -> Dict[str, Any]:
        """
        Reload the file from OneDrive (useful if external changes were made).
        
        Returns:
            Dictionary with success status
        """
        if not self.onedrive_path:
            return {
                "success": False,
                "message": "No file is currently open"
            }
        
        # Store path
        path = self.onedrive_path
        
        # Close current file
        self.close_file()
        
        # Reopen
        return self.open_file(path)
    
    def __repr__(self):
        """String representation of FileHandler."""
        if self.workbook:
            return f"FileHandler(file='{Path(self.onedrive_path).name}', sheets={len(self.sheet_names)}, read_only=True)"
        return "FileHandler(no file open)"
    
    def __del__(self):
        """Cleanup when FileHandler is destroyed."""
        if self.workbook:
            try:
                self.workbook.close()
            except:
                pass
