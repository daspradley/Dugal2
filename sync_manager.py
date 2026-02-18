"""
SyncManager - Orchestrates File Operations
Coordinates UniversalFileHandler (read) and TemporaryUpdateHandler (write).
Handles syncing changes between temp files and source (OneDrive, Google Sheets, etc.).
"""

import logging
import shutil
from pathlib import Path
from typing import Dict, Any, Optional
from datetime import datetime

logger = logging.getLogger(__name__)


class SyncManager:
    """
    Orchestrates file operations between read and write handlers.
    
    Architecture:
    - UniversalFileHandler: Read-only access to source files
    - TemporaryUpdateHandler: Write operations on temp copies
    - SyncManager: Coordinates both and handles syncing
    
    Workflow:
    1. open_file() → Creates both handlers, makes temp copy
    2. User makes updates → Goes to temp file
    3. save_to_source() → Copies temp back to source
    4. discard_changes() → Deletes temp, starts fresh
    5. reload_from_source() → Pulls latest from source if changed
    """
    
    def __init__(self, search_engine=None, google_credentials_path=None):
        """
        Initialize SyncManager.
        
        Args:
            search_engine: Search engine for item lookups
            google_credentials_path: Path to Google service account JSON (optional)
        """
        self.file_handler = None
        self.update_handler = None
        self.search_engine = search_engine
        self.google_credentials_path = google_credentials_path
        
        # Track source info
        self.source_path = None
        self.file_type = None
        self.is_open = False
        
        # Sync tracking
        self.last_sync_time = None
        self.source_modified_time = None
        
        logger.debug("SyncManager initialized")
    
    def open_file(self, file_path_or_url: str) -> Dict[str, Any]:
        """
        Open a file for reading and editing.
        Creates both read handler and temp write handler.
        
        Args:
            file_path_or_url: Path to Excel/CSV or Google Sheets URL
            
        Returns:
            Dictionary with success status and file info
        """
        try:
            # Import handlers
            from universal_file_handler import UniversalFileHandler
            from temporary_update_handler import TemporaryUpdateHandler
            
            # Create file handler (read-only)
            self.file_handler = UniversalFileHandler(
                search_engine=self.search_engine,
                google_credentials_path=self.google_credentials_path
            )
            
            # Open file for reading
            result = self.file_handler.open_file(file_path_or_url)
            
            if not result['success']:
                return result
            
            # Store source info
            self.source_path = self.file_handler.get_file_path()
            self.file_type = self.file_handler.get_file_type()
            
            # Track source modification time (for conflict detection)
            if self.file_type == "excel":
                self.source_modified_time = Path(self.source_path).stat().st_mtime
            
            # Create update handler (temp file for writes)
            self.update_handler = TemporaryUpdateHandler(
                source_path=self.source_path,
                file_type=self.file_type,
                file_handler=self.file_handler
            )
            
            self.is_open = True
            
            logger.info(f"✅ File opened successfully: {self.source_path}")
            logger.info(f"   Temp file: {self.update_handler.get_temp_path()}")
            
            return {
                "success": True,
                "message": "File opened successfully",
                "file_info": result['file_info'],
                "temp_path": str(self.update_handler.get_temp_path()),
                "source_path": self.source_path,
                "file_type": self.file_type
            }
            
        except Exception as e:
            error_msg = f"Error opening file: {str(e)}"
            logger.error(error_msg, exc_info=True)
            return {
                "success": False,
                "message": error_msg
            }
    
    def update_inventory(self, item_name: str, value: float, is_addition: bool = False) -> Dict[str, Any]:
        """
        Update inventory item (writes to temp file).
        
        Args:
            item_name: Name of the inventory item
            value: New value or amount to add/subtract
            is_addition: If True, add/subtract. If False, set to value.
            
        Returns:
            Dictionary with success status
        """
        if not self.is_open:
            return {
                "success": False,
                "message": "No file is open"
            }
        
        return self.update_handler.update_inventory(item_name, value, is_addition)
    
    def batch_update(self, updates: list) -> Dict[str, Any]:
        """
        Perform multiple inventory updates.
        
        Args:
            updates: List of update dictionaries
            
        Returns:
            Dictionary with batch results
        """
        if not self.is_open:
            return {
                "success": False,
                "message": "No file is open"
            }
        
        return self.update_handler.batch_update(updates)
    
    def save_to_source(self, backup: bool = True) -> Dict[str, Any]:
        """
        Save temp file back to source (OneDrive, Google Sheets, etc.).
        
        Args:
            backup: If True, create backup of source before overwriting
            
        Returns:
            Dictionary with success status
        """
        if not self.is_open:
            return {
                "success": False,
                "message": "No file is open"
            }
        
        if not self.update_handler.has_unsaved_changes():
            return {
                "success": True,
                "message": "No changes to save",
                "changes_saved": 0
            }
        
        try:
            temp_path = self.update_handler.get_temp_path()
            
            # Create backup if requested
            if backup and self.file_type == "excel":
                backup_path = self._create_backup()
                logger.info(f"Backup created: {backup_path}")
            
            # Save based on file type
            if self.file_type == "excel":
                # Excel: Direct file copy
                logger.info(f"Saving to Excel: {self.source_path}")
                shutil.copy(temp_path, self.source_path)
                
            elif self.file_type == "google_sheets":
                # Google Sheets: Upload temp file
                logger.info(f"Uploading to Google Sheets: {self.source_path}")
                self._upload_to_google_sheets(temp_path)
                
            elif self.file_type == "csv":
                # CSV: Convert and save
                logger.info(f"Saving to CSV: {self.source_path}")
                self._save_to_csv(temp_path)
            
            # Update sync time
            self.last_sync_time = datetime.now()
            
            # Update source modified time
            if self.file_type == "excel":
                self.source_modified_time = Path(self.source_path).stat().st_mtime
            
            # Get change summary
            summary = self.update_handler.get_change_summary()
            
            logger.info(f"✅ Saved {summary['total_changes']} changes to source")
            
            return {
                "success": True,
                "message": "Changes saved to source",
                "changes_saved": summary['total_changes'],
                "items_modified": summary['items_modified'],
                "last_sync": self.last_sync_time.isoformat()
            }
            
        except Exception as e:
            error_msg = f"Error saving to source: {str(e)}"
            logger.error(error_msg, exc_info=True)
            return {
                "success": False,
                "message": error_msg
            }
    
    def _create_backup(self) -> Path:
        """
        Create backup of source file.
        
        Returns:
            Path to backup file
        """
        source = Path(self.source_path)
        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        backup_name = f"{source.stem}_backup_{timestamp}{source.suffix}"
        backup_path = source.parent / backup_name
        
        shutil.copy(self.source_path, backup_path)
        return backup_path
    
    def _upload_to_google_sheets(self, temp_path: Path):
        """
        Upload temp Excel file back to Google Sheets.
        
        Args:
            temp_path: Path to temp Excel file
        """
        # TODO: Implement Google Sheets upload
        # This requires:
        # 1. Read Excel file
        # 2. Convert to Google Sheets format
        # 3. Use Google Sheets API to update
        raise NotImplementedError("Google Sheets upload coming in Phase 2!")
    
    def _save_to_csv(self, temp_path: Path):
        """
        Save temp Excel file as CSV.
        
        Args:
            temp_path: Path to temp Excel file
        """
        from openpyxl import load_workbook
        import csv
        
        wb = load_workbook(temp_path)
        ws = wb.active
        
        with open(self.source_path, 'w', newline='', encoding='utf-8') as f:
            writer = csv.writer(f)
            for row in ws.iter_rows(values_only=True):
                writer.writerow(row)
        
        logger.info(f"Saved to CSV: {self.source_path}")
    
    def discard_changes(self) -> Dict[str, Any]:
        """
        Discard all changes and recreate temp file from source.
        
        Returns:
            Dictionary with success status
        """
        if not self.is_open:
            return {
                "success": False,
                "message": "No file is open"
            }
        
        result = self.update_handler.discard_changes()
        
        if result['success']:
            logger.info("Changes discarded")
        
        return result
    
    def reload_from_source(self) -> Dict[str, Any]:
        """
        Check if source was modified externally and reload if needed.
        Useful for multi-user scenarios.
        
        Returns:
            Dictionary with reload status
        """
        if not self.is_open:
            return {
                "success": False,
                "message": "No file is open"
            }
        
        try:
            # Check if source has been modified
            if self.file_type == "excel":
                current_mtime = Path(self.source_path).stat().st_mtime
                
                if current_mtime <= self.source_modified_time:
                    return {
                        "success": True,
                        "message": "Source file has not changed",
                        "reloaded": False
                    }
                
                logger.info("Source file was modified externally - reloading")
            
            elif self.file_type == "google_sheets":
                logger.info("Reloading from Google Sheets")
            
            # Warn if unsaved changes
            if self.update_handler.has_unsaved_changes():
                logger.warning("⚠️ Reloading will discard unsaved changes!")
            
            # Reload file handler
            self.file_handler.reload_file()
            
            # Recreate update handler with fresh data
            self.update_handler.discard_changes()
            
            # Update modified time
            if self.file_type == "excel":
                self.source_modified_time = Path(self.source_path).stat().st_mtime
            
            return {
                "success": True,
                "message": "Reloaded from source",
                "reloaded": True
            }
            
        except Exception as e:
            error_msg = f"Error reloading from source: {str(e)}"
            logger.error(error_msg, exc_info=True)
            return {
                "success": False,
                "message": error_msg
            }
    
    def check_for_conflicts(self) -> Dict[str, Any]:
        """
        Check if source file was modified while we have unsaved changes.
        This detects potential conflicts.
        
        Returns:
            Dictionary with conflict status
        """
        if not self.is_open or self.file_type != "excel":
            return {
                "has_conflict": False,
                "message": "No conflict detection available"
            }
        
        # Check if we have unsaved changes
        has_changes = self.update_handler.has_unsaved_changes()
        
        # Check if source was modified
        current_mtime = Path(self.source_path).stat().st_mtime
        source_changed = current_mtime > self.source_modified_time
        
        # Conflict if both are true
        has_conflict = has_changes and source_changed
        
        if has_conflict:
            logger.warning("⚠️ CONFLICT DETECTED: Source file modified while you have unsaved changes!")
        
        return {
            "has_conflict": has_conflict,
            "unsaved_changes": has_changes,
            "source_modified": source_changed,
            "message": "Conflict detected!" if has_conflict else "No conflicts"
        }
    
    def get_status(self) -> Dict[str, Any]:
        """
        Get current sync manager status.
        
        Returns:
            Dictionary with status information
        """
        if not self.is_open:
            return {
                "is_open": False,
                "message": "No file is open"
            }
        
        return {
            "is_open": True,
            "source_path": self.source_path,
            "file_type": self.file_type,
            "temp_path": str(self.update_handler.get_temp_path()),
            "has_unsaved_changes": self.update_handler.has_unsaved_changes(),
            "change_summary": self.update_handler.get_change_summary(),
            "last_sync": self.last_sync_time.isoformat() if self.last_sync_time else None,
            "conflict_check": self.check_for_conflicts()
        }
    
    def close_file(self, save_changes: bool = False) -> Dict[str, Any]:
        """
        Close the file.
        
        Args:
            save_changes: If True, save changes before closing
            
        Returns:
            Dictionary with success status
        """
        if not self.is_open:
            return {
                "success": False,
                "message": "No file is open"
            }
        
        try:
            # Save if requested
            if save_changes and self.update_handler.has_unsaved_changes():
                save_result = self.save_to_source()
                if not save_result['success']:
                    return save_result
            
            # Close handlers
            if self.file_handler:
                self.file_handler.close_file()
            
            if self.update_handler:
                self.update_handler.close()
            
            # Reset state
            self.file_handler = None
            self.update_handler = None
            self.source_path = None
            self.file_type = None
            self.is_open = False
            
            logger.info("File closed")
            
            return {
                "success": True,
                "message": "File closed successfully"
            }
            
        except Exception as e:
            error_msg = f"Error closing file: {str(e)}"
            logger.error(error_msg, exc_info=True)
            return {
                "success": False,
                "message": error_msg
            }
    
    def __repr__(self):
        """String representation of SyncManager."""
        if self.is_open:
            filename = Path(self.source_path).name if self.file_type == "excel" else self.file_handler.file_info.get('filename', 'Unknown')
            changes = self.update_handler.get_change_summary()['total_changes'] if self.update_handler else 0
            return f"SyncManager(file='{filename}', type={self.file_type}, changes={changes}, open=True)"
        return "SyncManager(no file open)"
    
    def __enter__(self):
        """Context manager support."""
        return self
    
    def __exit__(self, exc_type, exc_val, exc_tb):
        """Context manager cleanup."""
        if self.is_open:
            self.close_file(save_changes=False)
