"""
Excel handler module for Dugal Inventory System.
Manages workbook viewing operations and refresh functionality.
"""

import os
import logging
import platform
import tempfile
import time
import unicodedata
import re
from difflib import SequenceMatcher, get_close_matches
from dataclasses import dataclass, field
from typing import List, Dict, Optional, Any, Tuple
from datetime import datetime

from openpyxl import load_workbook
from openpyxl.utils import column_index_from_string, get_column_letter
from PyQt5.QtWidgets import (
    QWidget, QVBoxLayout, QLabel, QPushButton, QCheckBox,
    QComboBox, QMessageBox, QDialog, QDialogButtonBox, QGroupBox,
    QScrollArea, QFrame
)
from PyQt5.QtCore import Qt, QTimer, pyqtSignal, QEventLoop
from PyQt5.QtGui import QIcon
from dictionary_manager import DictionaryManager
from search_engine import AdaptiveInventorySearchEngine

# Configure logging
logger = logging.getLogger(__name__)

@dataclass
class WorkbookState:
    """Tracks workbook viewing state and operations."""
    workbook: Optional[Any] = None
    dugal: Optional['MainDugal'] = None
    logging_manager: Optional[Any] = None
    file_path: Optional[str] = None
    temp_path: Optional[str] = None
    is_read_only: bool = True  # Default to read-only
    error_count: int = 0
    selected_sheets: List[str] = field(default_factory=list)
    search_column: Optional[str] = None
    input_column: Optional[str] = None
    last_refresh: Optional[datetime] = None
    refresh_interval: int = 2  # seconds
    active_sheet: Optional[str] = None
    view_status: Dict[str, Any] = field(default_factory=lambda: {
        'last_update': None,
        'refresh_count': 0,
        'view_errors': 0
    })

class ExcelControlPanel(QWidget):
    """Control panel for Excel viewing operations."""
    
    refresh_requested = pyqtSignal()
    
    def __init__(self, excel_handler, parent=None):
        super().__init__(parent)
        logger.debug("Initializing Excel Control Panel")
        self.excel_handler = excel_handler
        self.setup_ui()
        
        # Set up refresh timer
        self.refresh_timer = QTimer(self)
        self.refresh_timer.timeout.connect(self.request_refresh)
        
    def setup_ui(self):
        """Set up the control panel UI."""
        self.setWindowTitle("Dugal's Inventory Controls")
        self.setWindowFlags(Qt.Window | Qt.WindowStaysOnTopHint)
        self.setMinimumWidth(300)
        
        layout = QVBoxLayout(self)
        
        # Status section
        status_group = QGroupBox("Current Status")
        status_layout = QVBoxLayout()
        self.status_label = QLabel("Ready")
        self.file_label = QLabel("No file loaded")
        self.read_only_label = QLabel("Read-Only Mode")
        
        status_layout.addWidget(self.status_label)
        status_layout.addWidget(self.file_label)
        status_layout.addWidget(self.read_only_label)
        status_group.setLayout(status_layout)
        layout.addWidget(status_group)
        
        # Control buttons
        button_layout = QVBoxLayout()
        
        self.start_button = QPushButton("Start Updates")
        self.start_button.setStyleSheet("""
            QPushButton {
                background-color: #28a745;
                color: white;
                padding: 10px;
                border-radius: 5px;
                font-weight: bold;
            }
            QPushButton:hover {
                background-color: #218838;
            }
            QPushButton:disabled {
                background-color: #94d3a2;
            }
        """)
        
        self.pause_button = QPushButton("Pause Updates")
        self.pause_button.setStyleSheet("""
            QPushButton {
                background-color: #ffc107;
                color: black;
                padding: 10px;
                border-radius: 5px;
                font-weight: bold;
            }
            QPushButton:hover {
                background-color: #e0a800;
            }
            QPushButton:disabled {
                background-color: #ffe5a3;
            }
        """)
        self.pause_button.setEnabled(False)
        
        self.refresh_button = QPushButton("Refresh View")
        self.refresh_button.setStyleSheet("""
            QPushButton {
                background-color: #17a2b8;
                color: white;
                padding: 10px;
                border-radius: 5px;
                font-weight: bold;
            }
            QPushButton:hover {
                background-color: #138496;
            }
            QPushButton:disabled {
                background-color: #9ad0d9;
            }
        """)
        
        # Add Dictionary and scan buttons
        self.dict_button = QPushButton("Dictionary Manager")
        self.dict_button.setStyleSheet("""
            QPushButton {
                background-color: #6c5ce7;
                color: white;
                padding: 10px;
                border-radius: 5px;
                font-weight: bold;
            }
            QPushButton:hover {
                background-color: #5f51e8;
            }
        """)
        
        self.scan_button = QPushButton("Scan Document")
        self.scan_button.setStyleSheet("""
            QPushButton {
                background-color: #e67e22;
                color: white;
                padding: 10px;
                border-radius: 5px;
                font-weight: bold;
            }
            QPushButton:hover {
                background-color: #d35400;
            }
        """)
        
        self.end_button = QPushButton("End Session")
        self.end_button.setStyleSheet("""
            QPushButton {
                background-color: #dc3545;
                color: white;
                padding: 10px;
                border-radius: 5px;
                font-weight: bold;
            }
            QPushButton:hover {
                background-color: #c82333;
            }
        """)
        
        button_layout.addWidget(self.start_button)
        button_layout.addWidget(self.pause_button)
        button_layout.addWidget(self.refresh_button)
        button_layout.addWidget(self.dict_button)
        button_layout.addWidget(self.scan_button)
        button_layout.addWidget(self.end_button)
        layout.addLayout(button_layout)
        
        # Connect signals
        self.start_button.clicked.connect(self.start_updates)
        self.pause_button.clicked.connect(self.pause_updates)
        self.refresh_button.clicked.connect(self.request_refresh)
        self.dict_button.clicked.connect(self.show_dictionary)
        self.scan_button.clicked.connect(self.scan_document)
        self.end_button.clicked.connect(self.end_session)

        logger.debug("Control panel UI setup complete")

    def update_status(self, message: str):
        """Update the status display."""
        self.status_label.setText(message)
        
    def set_file_info(self, file_path: str):
        """Update the displayed file information."""
        file_name = os.path.basename(file_path)
        self.file_label.setText(f"Current file: {file_name}")
        
    def request_refresh(self):
        """Request a view refresh."""
        self.refresh_requested.emit()
        self.update_status("Refreshing view...")
        
    def start_updates(self):
        """Handle starting updates."""
        self.start_button.setEnabled(False)
        self.pause_button.setEnabled(True)
        self.update_status("Updates Active")
        if hasattr(self.excel_handler, 'start_updating'):
            self.excel_handler.start_updating()
        
    def pause_updates(self):
        """Handle pausing updates."""
        self.start_button.setEnabled(True)
        self.pause_button.setEnabled(False)
        self.update_status("Updates Paused")
        self.excel_handler.pause_updating()
        
    def end_session(self):
        """Handle ending the session."""
        reply = QMessageBox.question(
            self,
            'End Session',
            'Are you sure you want to end this session?',
            QMessageBox.Yes | QMessageBox.No,
            QMessageBox.No
        )
        
        if reply == QMessageBox.Yes:
            self.excel_handler.end_session()
            self.close()

    # New methods for dictionary and scanning functionality
    def show_dictionary(self):
        """Show the dictionary management interface."""
        logger.debug("Opening dictionary manager from control panel")
        try:
            self.excel_handler.show_dictionary_manager()
        except Exception as e:
            logger.error(f"Error opening dictionary manager: {e}")
            QMessageBox.warning(self, "Error", "Could not open dictionary manager")

    def scan_document(self):
        """Initiate document scan for terms."""
        logger.debug("Initiating document scan from control panel")
        try:
            if not self.excel_handler.state.workbook:
                logger.warning("No workbook loaded for scanning")
                QMessageBox.warning(self, "Error", "Please load a workbook first")
                return
                
            self.scan_button.setEnabled(False)
            self.update_status("Scanning document...")
            
            # Perform scan
            self.excel_handler.search_engine.learn_from_document(
                self.excel_handler.state.workbook,
                self.excel_handler.state.search_column
            )
            
            self.update_status("Scan complete")
            QMessageBox.information(self, "Success", "Document scan completed successfully!")
            
        except Exception as e:
            logger.error(f"Error during document scan: {e}")
            QMessageBox.warning(self, "Error", f"Error scanning document: {str(e)}")
        finally:
            self.scan_button.setEnabled(True)
            self.update_status("Ready")

class ColumnSelectionDialog(QDialog):
    """Dialog for selecting search and input columns."""
    
    def __init__(self, workbook, selected_sheets, parent=None, read_only=True):
        super().__init__(parent)
        logger.debug("Initializing Column Selection Dialog")
        self.workbook = workbook
        self.selected_sheets = selected_sheets
        self.search_column = None
        self.input_column = None
        self.read_only = read_only
        self.setWindowFlags(self.windowFlags()|Qt.WindowStaysOnTopHint)
        self.setup_ui()
        
    def setup_ui(self):
        """Set up the column selection dialog UI."""
        self.setWindowTitle("Column Selection")
        self.setMinimumWidth(400)
        layout = QVBoxLayout(self)
        
        if self.read_only:
            notice = QLabel("Read-Only Mode")
            notice.setStyleSheet("color: blue; font-weight: bold;")
            layout.addWidget(notice)
        
        # Search column selection (using Excel column letters)
        search_group = QGroupBox("Search Column")
        search_layout = QVBoxLayout()
        self.search_combo = QComboBox()
        
        # Add column letters A through M
        self.search_combo.addItems([get_column_letter(i) for i in range(1, 14)])

        #Add tooltip for search column
        search_help = QLabel("Select the column containing item names for advanced search")
        search_help.setStyleSheet("color: #666; font-style: italic;")
        search_layout.addWidget(QLabel("Select column letter to search in:"))
        search_layout.addWidget(search_help)
        search_layout.addWidget(self.search_combo)
        search_group.setLayout(search_layout)
        layout.addWidget(search_group)
        
        # Input column selection (using actual headers)
        input_group = QGroupBox("Input Column")
        input_layout = QVBoxLayout()
        self.input_combo = QComboBox()
        headers = self.get_all_headers()
        self.input_combo.addItems(headers)
        
        # Try to automatically select common inventory columns
        self._auto_select_inventory_column(headers)
        
        input_layout.addWidget(QLabel("Select column for quantity input:"))
        input_layout.addWidget(self.input_combo)
        input_group.setLayout(input_layout)
        layout.addWidget(input_group)
        
        # Preview section
        preview_group = QGroupBox("Selection Preview")
        preview_layout = QVBoxLayout()
        self.preview_label = QLabel()
        self.update_preview()
        preview_layout.addWidget(self.preview_label)
        preview_group.setLayout(preview_layout)
        layout.addWidget(preview_group)
        
        # Buttons
        button_box = QDialogButtonBox(
            QDialogButtonBox.Ok | QDialogButtonBox.Cancel
        )
        button_box.accepted.connect(self.accept)
        button_box.rejected.connect(self.reject)
        layout.addWidget(button_box)
        
        # Connect signals
        self.search_combo.currentTextChanged.connect(self.update_preview)
        self.input_combo.currentTextChanged.connect(self.update_preview)

    def validate_search_column(self, column: str) -> bool:
        """
        Validate that the selected search column contains searchable content.
        """
        try:
            column_idx = column_index_from_string(column)
            valid_content = False
            
            # Check content in each selected sheet
            for sheet_name in self.selected_sheets:
                sheet = self.workbook[sheet_name]
                non_empty_count = 0
                
                # Check first few rows for content
                for row in sheet.iter_rows(min_row=1, max_row=10, min_col=column_idx, max_col=column_idx):
                    if row[0].value and str(row[0].value).strip():
                        non_empty_count += 1
                
                if non_empty_count >= 3:  # At least 3 non-empty cells
                    valid_content = True
                    break
            
            return valid_content
            
        except Exception as e:
            logger.error(f"Error validating search column: {e}")
            return False

    def get_all_headers(self) -> List[str]:
        """Get unique column headers from all selected sheets."""
        headers = set()
        for sheet_name in self.selected_sheets:
            sheet = self.workbook[sheet_name]
            for idx, cell in enumerate(sheet[1], start=1):
                if idx > 1 and cell.value:
                    header_value = str(cell.value).strip()
                    if header_value:
                        headers.add(header_value)
        return sorted(list(headers))

    def _auto_select_inventory_column(self, headers: List[str]) -> None:
        """Try to automatically select common inventory column names."""
        inventory_keywords = ['quantity', 'qty', 'count', 'inventory', 'stock', 'amount']
        for idx, header in enumerate(headers):
            if any(keyword in header.lower() for keyword in inventory_keywords):
                self.input_combo.setCurrentIndex(idx)
                break
    
    def update_preview(self) -> None:
        """Update the preview of selected columns."""
        search_col = self.search_combo.currentText()
        input_col = self.input_combo.currentText()
        preview_text = (
            f"Will search in column {search_col} and update values in "
            f"'{input_col}' column"
        )
        if self.read_only:
            preview_text += " (Read-Only Mode)"
        self.preview_label.setText(preview_text)
        self.preview_label.setStyleSheet(
            "color: #2c3e50; padding: 5px; background-color: #ecf0f1; border-radius: 3px;"
        )
    
    def get_selections(self) -> Tuple[str, str]:
        """Return the selected column names."""
        return (
            self.search_combo.currentText(),
            self.input_combo.currentText()
        )
        
    # Modified accept method with enhanced validation
    def accept(self) -> None:
        """Validate selections before accepting."""
        selected_col = self.search_combo.currentText()
        logger.debug(f"Validating selected search column: {selected_col}")
        
        if not selected_col or not self.input_combo.currentText():
            logger.warning("Missing column selection")
            QMessageBox.warning(
                self,
                "Selection Required",
                "Please select both search and input columns."
            )
            return

        # Validate search column content
        if not self.validate_search_column(selected_col):
            logger.warning(f"Invalid search column content in column {selected_col}")
            reply = QMessageBox.question(
                self,
                "Column Validation",
                "The selected search column may not contain proper searchable content. "
                "This might affect search functionality. Continue anyway?",
                QMessageBox.Yes | QMessageBox.No,
                QMessageBox.No
            )
            if reply == QMessageBox.No:
                return

        logger.debug(f"Column selections validated - Search: {selected_col}, "
                    f"Input: {self.input_combo.currentText()}")
        super().accept()

class ExcelHandler(QWidget):
    """Handles Excel file viewing operations."""

    refresh_complete = pyqtSignal()

    def __init__(self, dugal=None):
        super().__init__()
        logger.debug("Initializing Excel handler")
        self.state = WorkbookState(dugal=dugal)
        self.control_panel = None
        self.is_updating = False
        self.sheet_checkboxes = {}
        
        # Get logging manager from dugal if available
        self.logging_manager = dugal.logging_manager if dugal and hasattr(dugal, 'logging_manager') else None

        # Initialize search engine and dictionary manager
        self.search_engine = AdaptiveInventorySearchEngine()
        self.dictionary_manager = None

        self.setup_ui()
        logger.debug("Excel handler initialized with Dugal's personality")

    def setup_ui(self) -> None:
        """Set up the user interface for sheet selection."""
        self.setWindowTitle("Dugal's Sheet Selector")
        self.setGeometry(300, 300, 400, 400)
        
        layout = QVBoxLayout(self)
        
        # Sheet selection header
        header_label = QLabel("Select sheets to parse:")
        header_label.setStyleSheet("font-weight: bold;")
        layout.addWidget(header_label)

        # Create scrollable area for sheets
        scroll_area = QScrollArea(self)
        scroll_area.setWidgetResizable(True)
        scroll_content = QWidget()
        self.sheet_layout = QVBoxLayout(scroll_content)
        
        # Select All checkbox
        self.select_all_checkbox = QCheckBox("Select All")
        self.select_all_checkbox.stateChanged.connect(self.select_all_sheets)
        self.sheet_layout.addWidget(self.select_all_checkbox)
        
        # Add line separator
        line = QFrame()
        line.setFrameShape(QFrame.HLine)
        line.setFrameShadow(QFrame.Sunken)
        self.sheet_layout.addWidget(line)
        
        scroll_area.setWidget(scroll_content)
        layout.addWidget(scroll_area)
        
        # Proceed button
        self.proceed_button = QPushButton("Proceed")
        self.proceed_button.setStyleSheet("""
            QPushButton {
                background-color: #2c3e50;
                color: white;
                padding: 8px;
                border-radius: 4px;
                font-weight: bold;
            }
            QPushButton:hover {
                background-color: #34495e;
            }
        """)
        self.proceed_button.clicked.connect(self.apply_selections)
        layout.addWidget(self.proceed_button)

        # Add Dictionary Manager button
        #self.dict_button = QPushButton("Dictionary Manager")
        #self.dict_button.setStyleSheet("""
            #QPushButton {
             #   background-color: #6c5ce7;
              #  color: white;
               # padding: 8px;
                #border-radius: 4px;
                #font-weight: bold;
            #}
            #QPushButton:hover {
                #background-color: #5f51e8;
            #}
        #""")
        #self.dict_button.clicked.connect(self.show_dictionary_manager)
        #layout.addWidget(self.dict_button)
        #logger.debug("Dictionary manager button added to UI")

    def _show_sheet_dialog(self, workbook):
        """Show sheet selection dialog with proper focus."""
        # For modal behavior, we're already using exec_() in the load_excel_file method
        # Make sure window has proper focus
        self.setWindowFlags(self.windowFlags() | Qt.WindowStaysOnTopHint)
        self.raise_()
        self.activateWindow()

    def show_dictionary_manager(self):
        """Show the dictionary management interface."""
        logger.debug("Opening dictionary manager")
        try:
            if not self.dictionary_manager:
                # Create dictionary manager with all required components
                self.dictionary_manager = DictionaryManager(
                    search_engine = self.search_engine,
                    parent = self,
                    logging_manager=self.logging_manager,
                    workbook_state=self.state
                )
                # Connect signals for logging and updates
                self.dictionary_manager.term_added.connect(self._handle_new_term)
                self.dictionary_manager.term_updated.connect(self._handle_term_update)

            self.dictionary_manager.show()
            self.dictionary_manager.raise_()
            self.dictionary_manager.activateWindow()

            if self.logging_manager:
                self.logging_manager.log_pattern_match({
                    'type': 'dictionary_manager_opened',
                    'timestamp': datetime.now().isoformat()
                })
                
        except Exception as e:
            logger.error(f"Error opening dictionary manager: {e}")
            if self.logging_manager:
                self.logging_manager.log_error(str(e), {
                    'context': 'dictionary_manager',
                    'action': 'open'
                })
            QMessageBox.warning(self, "Error", "Could not open dictionary manager")

    def _handle_new_term(self, term_data: Dict[str, Any]):
        """Handle newly added dictionary terms."""
        try:
            # Update search engine with new term
            self.search_engine.add_term_with_variations(
                term_data['base_term'],
                term_data['variations']
            )
            
            if self.logging_manager:
                self.logging_manager.log_pattern_match({
                    'type': 'new_term_added',
                    'term': term_data,
                    'timestamp': datetime.now().isoformat()
                })
                
        except Exception as e:
            logger.error(f"Error handling new term: {e}")
            if self.logging_manager:
                self.logging_manager.log_error(str(e), {
                    'context': 'dictionary_manager',
                    'action': 'add_term'
                })

    def _handle_term_update(self, update_data: Dict[str, Any]):
        """Handle dictionary term updates."""
        try:
            # Update search engine with modified term
            self.search_engine.update_term(
                update_data['old_term'],
                update_data['new_term'],
                update_data['variations']
            )
            
            if self.logging_manager:
                self.logging_manager.log_pattern_match({
                    'type': 'term_updated',
                    'update': update_data,
                    'timestamp': datetime.now().isoformat()
                })
                
        except Exception as e:
            logger.error(f"Error handling term update: {e}")
            if self.logging_manager:
                self.logging_manager.log_error(str(e), {
                    'context': 'dictionary_manager',
                    'action': 'update_term'
                })

    def load_excel_file(self, file_path: str, read_only: bool = True) -> bool:
        """Load Excel file and handle sheet/column selection."""
        try:
            logger.debug(f"Loading Excel file at {file_path} (read_only={read_only})")
            
            # Check file existence
            if not os.path.exists(file_path):
                logger.error("File not found: %s", file_path)
                return False

            # Load workbook
            self.state.workbook = load_workbook(file_path, read_only=read_only, data_only=True)
            self.state.file_path = file_path
            self.state.is_read_only = read_only
            
            # Initialize search engine with new file
            logger.debug("Initializing search engine with new file")
            self.search_engine = AdaptiveInventorySearchEngine()
            
            # Clear existing selections
            self.clear_existing_selections()
            
            # Show sheet selection
            self.populate_sheets()
            
            # Make window modal and ensure it's visible on top
            self.setWindowModality(Qt.ApplicationModal)
            self.setWindowFlags(self.windowFlags() | Qt.WindowStaysOnTopHint)
            self.show()
            self.raise_()
            self.activateWindow()
            
            logger.debug("Waiting for sheet selection...")
            if not self.exec_():
                logger.debug("Sheet selection cancelled")
                return False
            
            # After sheets selected, show column selection dialog
            if self.state.selected_sheets:
                column_dialog = ColumnSelectionDialog(
                    self.state.workbook,
                    self.state.selected_sheets,
                    self,
                    read_only=read_only
                )
                if column_dialog.exec_() != QDialog.Accepted:
                    logger.debug("Column selection cancelled")
                    return False
                
                # Get column selections
                self.state.search_column, self.state.input_column = column_dialog.get_selections()
                
                # Build search index now that we have the search column
                logger.debug("Building initial search index")
                self.build_search_index()
                
                # Open Excel and show control panel if in read-only mode
                if read_only:
                    self._open_excel_file(file_path)
                
                self._show_control_panel()
                return True
            
            logger.debug("No sheets selected")
            return False
            
        except Exception as e:
            logger.error(f"Error loading Excel file: {e}")
            return False

    def populate_sheets(self) -> None:
        """Populate checkboxes with sheet names."""
        for sheet_name in self.state.workbook.sheetnames:
            checkbox = QCheckBox(sheet_name)
            checkbox.stateChanged.connect(self.update_sheet_selection)
            self.sheet_checkboxes[sheet_name] = checkbox
            self.sheet_layout.addWidget(checkbox)

    def select_all_sheets(self, state: int) -> None:
        """Toggle all sheet checkboxes."""
        select_all = state == Qt.Checked
        for checkbox in self.sheet_checkboxes.values():
            checkbox.setChecked(select_all)

    def update_sheet_selection(self) -> None:
        """Update selected sheets list."""
        self.state.selected_sheets = [
            name for name, checkbox in self.sheet_checkboxes.items() 
            if checkbox.isChecked()
        ]
        
        all_checked = all(checkbox.isChecked() for checkbox in self.sheet_checkboxes.values())
        any_unchecked = any(not checkbox.isChecked() for checkbox in self.sheet_checkboxes.values())
        
        if all_checked:
            self.select_all_checkbox.setCheckState(Qt.Checked)
        elif any_unchecked:
            self.select_all_checkbox.setCheckState(Qt.Unchecked)

    def clear_existing_selections(self):
        """Clear existing sheet selections."""
        if hasattr(self, 'sheet_checkboxes'):
            for checkbox in self.sheet_checkboxes.values():
                self.layout().removeWidget(checkbox)
                checkbox.deleteLater()
            self.sheet_checkboxes.clear()

        if hasattr(self, 'select_all_checkbox'):
            self.select_all_checkbox.setChecked(False)

    def apply_selections(self) -> None:
        """Apply selected sheets and close dialog."""
        if not self.state.selected_sheets:
            QMessageBox.warning(self, "Selection Error", "Please select at least one sheet.")
            return
        self.close()

    def exec_(self) -> bool:
        """Make window modal and return selection status."""
        self.setWindowModality(Qt.ApplicationModal)
        self.show()
        loop = QEventLoop()
        self.finished = False
        
        def on_close():
            self.finished = bool(self.state.selected_sheets)
            loop.quit()
        
        self.proceed_button.clicked.connect(on_close)
        loop.exec_()
        return self.finished

    def _open_excel_file(self, file_path: str) -> None:
        """Open file in Excel application."""
        try:
            if not os.path.exists(file_path):
                logger.error("Failed to open Excel file: File does not exist")
                return

            if platform.system() == 'Windows':
                os.startfile(file_path)
                time.sleep(1)  # Give Excel time to open
                logger.debug("File opened in Excel (Windows)")
            else:
                subprocess.run(['xdg-open' if platform.system() == 'Linux' else 'open', file_path], check=True)
                logger.debug(f"File opened in Excel ({platform.system()})")
                
        except Exception as e:
            logger.error(f"Failed to open Excel file: {e}")

    def _show_control_panel(self):
        """Show the control panel after successful file loading."""
        try:
            if not hasattr(self, 'control_panel') or not self.control_panel:
                self.control_panel = ExcelControlPanel(self)
            
            if self.state.file_path:
                self.control_panel.set_file_info(self.state.file_path)
            
            if self.state.is_read_only:
                if hasattr(self.control_panel, 'refresh_timer'):
                    self.control_panel.refresh_timer.start(self.state.refresh_interval * 1000)
            
            self.control_panel.show()
            self.control_panel.raise_()
            self.control_panel.activateWindow()
            logger.debug("Control panel displayed and configured")
            
        except Exception as e:
            logger.error(f"Error showing control panel: {e}")

    def start_updating(self):
        """Start accepting updates."""
        self.is_updating = True
        if self.state.dugal:
            self.state.dugal.speak("Ready for inventory updates.")
            try:
                if hasattr(self.state.dugal, 'voice_interaction'):
                    self.state.dugal.voice_interaction.start_listening()
                else:
                    logger.error("Voice interaction not properly connected")
            except Exception as e:
                logger.error(f"Error starting voice updates: {e}")

    def build_search_index(self, workbook=None, search_column=None):
        """Build search index from workbook."""
        logger.debug(f"Building initial search index (search engine id: {id(self.search_engine)})")
        
        # Get search engine from global registry first
        try:
            # Try to use component manager first
            try:
                from component_manager import component_manager
                
                # Get search engine from component manager
                cm_engine = component_manager.get_search_engine()
                
                if cm_engine:
                    if id(cm_engine) != id(self.search_engine):
                        # Update our reference to match component manager
                        logger.debug(f"Updating search engine reference from component manager (old: {id(self.search_engine)}, new: {id(cm_engine)})")
                        self.search_engine = cm_engine
                else:
                    # Register our search engine in component manager
                    component_manager.register_component('search_engine', self.search_engine)
                    logger.debug(f"Registered search engine in component manager: {id(self.search_engine)}")
                    
            except ImportError:
                logger.debug("Component manager not available, falling back to legacy method")
                
                # Fall back to legacy method if component manager is not available
                from global_registry import GlobalRegistry
                registry_engine = GlobalRegistry.get('search_engine')
                
                # If not in registry, use local reference or register current one
                if not registry_engine:
                    GlobalRegistry.register('search_engine', self.search_engine)
                elif id(registry_engine) != id(self.search_engine):
                    # Update our reference to match registry
                    logger.debug(f"Updating search engine reference from registry (old: {id(self.search_engine)}, new: {id(registry_engine)})")
                    self.search_engine = registry_engine
        except Exception as e:
            logger.error(f"Error checking component manager/registry: {e}")
        
        # Use provided workbook or the current one
        wb = workbook if workbook else self.state.workbook
        col = search_column if search_column else self.state.search_column
        
        if not wb:
            logger.error("No workbook available for building search index")
            return False
            
        # Pass the workbook to the search engine for value lookup
        self.search_engine.workbook = wb
        logger.debug(f"Assigned workbook to search engine: {bool(wb)}")
        
        # Set the input column index for value lookup
        input_column_found = False
        if self.state.input_column:
            # Find the column index for the input column
            for sheet_name in self.state.selected_sheets:
                sheet = wb[sheet_name]
                for idx, cell in enumerate(sheet[1], start=1):
                    if cell.value and str(cell.value).strip() == self.state.input_column:
                        self.search_engine.input_column_index = idx
                        self.search_engine.input_column_name = self.state.input_column
                        logger.debug(f"Found input column '{self.state.input_column}' at index {idx}")
                        input_column_found = True
                        break
                if input_column_found:
                    break
                    
        if not input_column_found:
            logger.warning(f"Could not find input column '{self.state.input_column}' in header row")
            
        # Log the state before building index
        logger.debug(f"Search engine ready with: workbook={bool(wb)}, input_column_index={getattr(self.search_engine, 'input_column_index', None)}")
        
        # Build the index
        self.search_engine.build_search_index(wb, col)
        
        # After building, update the global registry
        try:
            from global_registry import GlobalRegistry
            GlobalRegistry.register('search_engine', self.search_engine)
            logger.debug(f"Updated global registry with current search engine (ID: {id(self.search_engine)})")
        except Exception as e:
            logger.error(f"Error updating global registry: {e}")
        
        # Connect voice interaction to search engine
        if self.state.dugal and hasattr(self.state.dugal, 'voice_interaction'):
            logger.debug(f"Connecting voice interaction directly to search engine (search engine id: {id(self.search_engine)})")
            
            # Use connect_to_search_engine method if available
            if hasattr(self.state.dugal.voice_interaction, 'connect_to_search_engine'):
                self.state.dugal.voice_interaction.connect_to_search_engine(self.search_engine)
            else:
                # Direct assignment as fallback
                self.state.dugal.voice_interaction.state.search_engine = self.search_engine
            
            # Also set in the global config for good measure
            if hasattr(self.state, 'config') and hasattr(self.state.config, 'set_global'):
                self.state.config.set_global('search_engine', self.search_engine)
                logger.debug("Updated global config with search engine reference")
                
            logger.debug(f"Connected voice interaction to search engine with {len(self.search_engine.inventory_cache) if hasattr(self.search_engine, 'inventory_cache') else 0} items")
            
            # Force voice interaction to diagnose its search engine connection
            if hasattr(self.state.dugal.voice_interaction, 'diagnose_search_engine'):
                self.state.dugal.voice_interaction.diagnose_search_engine()
                
        return True

    def get_search_engine(self):
        """Get the search engine from the global registry or local reference."""
        try:
            # First try global registry
            from global_registry import GlobalRegistry
            search_engine = GlobalRegistry.get('search_engine')
            
            # If found in registry, update our local reference
            if search_engine:
                if hasattr(self, 'search_engine') and self.search_engine is not search_engine:
                    logger.debug(f"Updating local search engine reference from registry")
                    self.search_engine = search_engine
                return search_engine
                
            # If not in registry but we have one, register it
            if hasattr(self, 'search_engine') and self.search_engine:
                GlobalRegistry.register('search_engine', self.search_engine)
                return self.search_engine
                
            # Last resort: if excel_handler has one
            if hasattr(self, 'excel_handler') and hasattr(self.excel_handler, 'search_engine'):
                search_engine = self.excel_handler.search_engine
                GlobalRegistry.register('search_engine', search_engine)
                return search_engine
                
            logger.warning("No search engine available")
            return None
            
        except Exception as e:
            logger.error(f"Error getting search engine: {e}")
            # Fallback to local reference if available
            return getattr(self, 'search_engine', None)

    def search_inventory(self, sheet, search_term: str, column_index: int = 1) -> Optional[int]:
        """
        Enhanced search for an item using NLP capabilities.
        """
        logger.debug(f"Starting enhanced search for '{search_term}' in sheet '{sheet.title}'")
        
        try:
            # Get search engine using the accessor method
            search_engine = self.get_search_engine()
            if not search_engine:
                logger.error("No search engine available")
                return None
            
            # Ensure search engine is initialized with current data
            if not search_engine.has_index():
                logger.debug("Building search index")
                search_engine.build_search_index(
                    self.state.workbook,
                    self.state.search_column
                )
            
            # Perform search
            match_row = search_engine.search(sheet, search_term)
            
            if match_row:
                logger.debug(f"Found match in row {match_row}")
                return match_row
            
            logger.debug(f"No match found for '{search_term}'")
            return None
            
        except Exception as e:
            logger.error(f"Error during inventory search: {e}")
            return None

    def check_column_content(self, sheet, column_index: int, term: str) -> bool:
        """
        Check if a column contains a specific term.
        Useful for verifying search column selection.
        """
        logger.debug(f"Checking column {column_index} for term: {term}")
        try:
            for row in sheet.iter_rows(min_row=1, min_col=column_index, max_col=column_index):
                cell_value = str(row[0].value or "").strip().lower()
                if term.lower() in cell_value:
                    return True
            return False
        except Exception as e:
            logger.error(f"Error checking column content: {e}")
            return False

    def validate_search_column(self, sheet, column_index: int) -> bool:
        """
        Validate that the selected column is appropriate for searching.
        """
        logger.debug(f"Validating search column {column_index}")
        try:
            # Check first few rows for non-empty values
            non_empty_count = 0
            for row in sheet.iter_rows(min_row=1, max_row=10, min_col=column_index, max_col=column_index):
                if row[0].value:
                    non_empty_count += 1
            
            # Column should have at least 3 non-empty values in first 10 rows
            return non_empty_count >= 3
        except Exception as e:
            logger.error(f"Error validating search column: {e}")
            return False

    def pause_updating(self):
        """Pause updates."""
        self.is_updating = False
        if self.state.dugal:
            self.state.dugal.speak("Updates paused.")

    def refresh_view(self) -> bool:
        """Refresh Excel view with latest data."""
        if not self.state.is_read_only:
            return False
            
        try:
            logger.debug("Starting view refresh...")
            current_time = datetime.now()
            if (self.state.last_refresh and
                (current_time - self.state.last_refresh).total_seconds() 
                < self.state.refresh_interval):
                logger.debug("Skipping refresh - too soon since last refresh")
                return True

            # Load latest workbook
            logger.debug("Loading latest workbook...")
            self.state.workbook = load_workbook(
                self.state.file_path,
                data_only=True,
                read_only=True
            )
            logger.debug("Workbook reloaded successfully.")

            # Update each sheet
            for sheet_name in self.state.selected_sheets:
                try:
                    worksheet = self.state.workbook[sheet_name]
                    worksheet_index = self.state.workbook.sheetnames.index(sheet_name) + 1
                    self.state.workbook.active = worksheet_index
                    logger.debug(f"Refreshed sheet: {sheet_name}")
                except Exception as sheet_error:
                    logger.error(f"Error refreshing sheet {sheet_name}: {sheet_error}")
                    continue

            self.state.last_refresh = current_time
            self.state.view_status['refresh_count'] += 1
            
            if self.control_panel:
                self.control_panel.update_status("View refreshed")
                
            return True
            
        except Exception as e:
            logger.error(f"Error refreshing view: {e}")
            self.state.view_status['view_errors'] += 1
            
            # Attempt recovery
            try:
                logger.debug("Attempting recovery after view refresh failure")
                # Try to reload the workbook with different options
                if self._attempt_view_recovery():
                    logger.info("Successfully recovered from view refresh failure")
                    return True
            except Exception as recovery_error:
                logger.error(f"Recovery attempt failed: {recovery_error}")
                
            return False

    def _attempt_view_recovery(self) -> bool:
        """Attempt to recover from view refresh failures."""
        try:
            logger.debug("Attempting view recovery")
            
            # Check if file exists
            if not os.path.exists(self.state.file_path):
                logger.warning("Excel file not found, attempting to locate")
                # Try to find the file in common locations
                common_locations = [
                    os.path.join(os.path.expanduser("~"), "Desktop"),
                    os.path.join(os.path.expanduser("~"), "OneDrive", "Desktop"),
                    os.path.join(os.path.expanduser("~"), "Documents")
                ]
                
                filename = os.path.basename(self.state.file_path)
                for location in common_locations:
                    potential_path = os.path.join(location, filename)
                    if os.path.exists(potential_path):
                        logger.info(f"Found Excel file at {potential_path}")
                        self.state.file_path = potential_path
                        break
                else:
                    return False
            
            # Try to load with different options
            try:
                logger.debug("Attempting to load workbook with alternative options")
                self.state.workbook = load_workbook(
                    self.state.file_path,
                    data_only=True,
                    read_only=False  # Try without read-only mode
                )
                logger.debug("Workbook loaded with alternative options")
                return True
            except Exception as alt_load_error:
                logger.error(f"Alternative load failed: {alt_load_error}")
                return False
                
        except Exception as e:
            logger.error(f"Error in view recovery: {e}")
            return False
    
    def end_session(self):
        """End the session."""
        try:
            if self.state.dugal:
                self.state.dugal.speak("Session ended. All updates saved.")
            self.cleanup()
        except Exception as e:
            logger.error(f"Error ending session: {e}")

    def cleanup(self) -> None:
        """Clean up resources before shutdown."""
        try:
            # Stop timers
            if self.control_panel and hasattr(self.control_panel, 'refresh_timer'):
                self.control_panel.refresh_timer.stop()
            
            # Close control panel
            if self.control_panel:
                self.control_panel.close()
                
            # Close workbook
            if self.state.workbook:
                try:
                  self.state.workbook.close()
                except Exception as e:
                    logger.error(f"Error closing workbook: {e}")
            
            # Reset state
            self.state = WorkbookState()
            logger.debug("Excel handler cleanup completed")
            
        except Exception as e:
              logger.error(f"Error during Excel handler cleanup: {e}")
