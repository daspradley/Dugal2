"""
OneDrive handler module for Dugal Inventory System.
Handles file operations, synchronization, and inventory updates with OneDrive.
"""

from __future__ import annotations

# Standard library imports
import os
import time
import shutil
import re
import subprocess
import json
from difflib import SequenceMatcher, get_close_matches
from dataclasses import dataclass, field, InitVar
from datetime import datetime
from typing import Dict, List, Optional, Any, Tuple, TYPE_CHECKING
from global_registry import GlobalRegistry

# Configure module logger at the top level
import logging
logger = logging.getLogger(__name__)

# Type hints for forward references
if TYPE_CHECKING:
    from main_dugal import MainDugal
    from logging_manager import LoggingManager
    from search_engine import AdaptiveInventorySearchEngine
    from data_manager import DataManager

# Third-party imports
from openpyxl import load_workbook
from openpyxl.utils import column_index_from_string
from PyQt5.QtCore import QObject, pyqtSignal

@dataclass
class OneDriveState:
    """Tracks OneDrive file state and operations."""
    dugal: Optional['MainDugal'] = None
    logging_manager: Optional['LoggingManager'] = None
    search_engine: Optional['AdaptiveInventorySearchEngine'] = None
    local_file_path: Optional[str] = None
    last_sync_time: Optional[datetime] = None
    is_file_locked: bool = False
    sync_interval: int = 30  # Minimum seconds between syncs
    file_access_attempts: int = 0
    sync_failures: int = 0
    lock_holder: Optional[str] = None
    lock_time: Optional[datetime] = None
    read_only: bool = False
    selected_sheets: List[str] = field(default_factory=list)
    search_column: Optional[str] = None
    input_column: Optional[str] = None
    status: Dict[str, Any] = field(default_factory=lambda: {
        'last_modified': None,
        'sync_status': 'unknown',
        'access_count': 0,
        'error_count': 0,
        'last_backup': None
    })
    recent_operations: List[Dict] = field(default_factory=list)
    pending_updates: List[Dict] = field(default_factory=list)
    cache: Dict[str, Any] = field(default_factory=lambda: {
        'items': {},
        'values': {},
        'matches': {}
    })

class OneDriveHandler(QObject):
    file_updated = pyqtSignal()
    sync_complete = pyqtSignal()

    def __init__(self, dugal=None):
        """Initialize the OneDrive handler."""
        super().__init__()
        # Initialize instance logger
        self.logger = logging.getLogger(__name__)
        self.logger.debug("Initializing OneDrive handler - Start")
        
        try:
            self.logger.debug("Setting up initial state")
            # Initialize state first with clean defaults
            self.state = OneDriveState(
                dugal=dugal,
                last_sync_time=None,
                is_file_locked=False,
                read_only=False,
                sync_failures=0,
                file_access_attempts=0
            )
            self.logger.debug("Initial state created successfully")
            
            # Create backup directory before any file operations
            self.backup_dir = ".dugal_backups"
            self.logger.debug(f"Creating backup directory at {self.backup_dir}")
            os.makedirs(self.backup_dir, exist_ok=True)
            
            # Initialize cache dictionary
            self.logger.debug("Initializing cache dictionary")
            self.state.cache = {
                'items': {},
                'values': {},
                'matches': {},
                'columns': {}
            }
            
            # Set up components from dugal if available
            self.logger.debug("Setting up components from dugal")
            if dugal:
                self.logger.debug("Dugal object provided, getting components")
                self.data_manager = getattr(dugal, 'data_manager', None)
                self.logging_manager = getattr(dugal, 'logging_manager', None)
                self.search_engine = getattr(dugal.state, 'search_engine', None) if hasattr(dugal, 'state') else None
                self.excel_handler = getattr(dugal, 'excel_handler', None)
                self.logger.debug(f"Components retrieved - data_manager: {bool(self.data_manager)}, logging_manager: {bool(self.logging_manager)}, search_engine: {bool(self.search_engine)}, excel_handler: {bool(self.excel_handler)}")
            else:
                self.logger.debug("No dugal object provided, initializing with None components")
                self.data_manager = None
                self.logging_manager = None
                self.get_search_engine = None
                self.excel_handler = None
            
            self.max_operations_history = 50
            
            self.logger.info("OneDrive handler initialized successfully")
            
        except Exception as e:
            self.logger.error(f"Error during OneDrive handler initialization: {e}", exc_info=True)
            raise RuntimeError(f"OneDrive handler initialization failed: {e}")

    def _ensure_backup_dir(self) -> None:
        """Ensure backup directory exists."""
        if not os.path.exists(self.backup_dir):
            os.makedirs(self.backup_dir)

    def get_synced_path(self, file_path: str) -> Optional[str]:
        """Get the locally synced path for a OneDrive file."""
        try:
            if self.ensure_local_copy(file_path):
                self.state.local_file_path = file_path
                self.logger.debug("File synced at: %s", file_path)
                return file_path
            return None
        except Exception as e:
            self.logger.error("Error getting synced path: %s", str(e))
            return None

    def ensure_local_copy(self, file_path: str) -> bool:
        """Ensure the specified OneDrive file is available locally."""
        self.logger.debug("Ensuring OneDrive file is available at: %s", file_path)
        try:
            self.state.file_access_attempts += 1
            
            if not os.path.exists(file_path):
                error_msg = self._get_dugal_response('error', "File not found")
                self.logger.error(error_msg)
                self._log_operation('file_access', False, "File not found")
                return False

            # Check if file is locked first
            if self.check_file_locks(file_path):
                error_msg = "File is locked by another process"
                self.logger.warning(error_msg)
                self._log_operation('file_access', False, error_msg)
                return False

            # Different handling for read-only mode
            if self.state.read_only:
                if not self.is_read_only_accessible(file_path):
                    return False
            else:
                # Verify full file access for non-read-only mode
                with open(file_path, 'rb+') as f:
                    f.read(1)

            # Update state
            self.state.local_file_path = file_path
            self.state.last_sync_time = datetime.now()
            self.state.status.update({
                'last_modified': os.path.getmtime(file_path),
                'access_count': self.state.status['access_count'] + 1,
                'sync_status': 'synced'
            })
            
            self._log_operation('file_access', True)
            success_msg = self._get_dugal_response('success', "File found and ready")
            self.logger.debug(success_msg)
            return True

        except Exception as e:
            error_msg = self._get_dugal_response('error', str(e))
            self._log_operation('file_access', False, str(e))
            self.logger.error(error_msg)
            return False

    def remove_excel_lock_files(self, file_path):
        """
        NUCLEAR OPTION: Aggressively delete Excel temporary lock files.
        
        Excel creates lock files like:
        - ~$filename.xlsx
        - .~lock.filename.xlsx#
        
        This method finds and deletes them to force-unlock the file.
        """
        import os
        
        if not file_path or not os.path.exists(file_path):
            return
        
        directory = os.path.dirname(file_path)
        filename = os.path.basename(file_path)
        
        # Common Excel lock file patterns
        lock_patterns = [
            f"~${filename}",  # Standard Excel lock
            f".~lock.{filename}#",  # LibreOffice/OpenOffice lock
        ]
        
        deleted_files = []
        
        for pattern in lock_patterns:
            lock_file_path = os.path.join(directory, pattern)
            if os.path.exists(lock_file_path):
                try:
                    os.remove(lock_file_path)
                    deleted_files.append(pattern)
                    self.logger.info(f"🔥 DELETED PHANTOM LOCK FILE: {pattern}")
                except Exception as e:
                    self.logger.warning(f"Failed to delete lock file {pattern}: {e}")
        
        if deleted_files:
            self.logger.info(f"Removed {len(deleted_files)} lock file(s): {deleted_files}")
            if self.logging_manager:
                self.logging_manager.log_pattern_match({
                    'type': 'phantom_locks_removed',
                    'files': deleted_files,
                    'timestamp': datetime.now().isoformat()
                })
        else:
            self.logger.debug("No phantom lock files found")

    def lock_file(self, max_retries=3, retry_delay=1.0) -> bool:
        """Lock file for editing with proper file system locking and retries."""
        import time
        
        if not self.state.local_file_path:
            if self.logging_manager:
                self.logging_manager.log_error("No file selected", {
                    'context': 'file_lock',
                    'action': 'lock'
                })
            self.logger.error("No file selected for locking")
            return False

        # Skip locking in read-only mode
        if self.state.read_only:
            self.logger.debug("Read-only mode, skipping file lock")
            if self.logging_manager:
                self.logging_manager.log_pattern_match({
                    'type': 'lock_skip',
                    'reason': 'read_only_mode',
                    'timestamp': datetime.now().isoformat()
                })
            return True

        # NUCLEAR OPTION: Delete phantom Excel lock files
        self.remove_excel_lock_files(self.state.local_file_path)
        
        for attempt in range(max_retries):
            try:
                lock_path = f"{self.state.local_file_path}.lock"
                
                # First re-check if the file is locked by any process
                if self.check_file_locks(self.state.local_file_path):
                    if attempt < max_retries - 1:
                        self.logger.warning(f"File locked by another process, retrying in {retry_delay} seconds (attempt {attempt+1}/{max_retries})")
                        time.sleep(retry_delay)
                        continue
                    else:
                        self.logger.error(f"File still locked after {max_retries} attempts")
                        if self.logging_manager:
                            self.logging_manager.log_error("File locked by another process", {
                                'context': 'file_lock',
                                'action': 'lock',
                                'attempts': max_retries
                            })
                        return False
                        
                # Check if lock exists
                if os.path.exists(lock_path):
                    try:
                        with open(lock_path, 'r', encoding='utf-8') as f:
                            existing_lock = eval(f.read())
                        
                        # Check if it's our lock
                        if (existing_lock.get('holder') == 'Dugal Inventory System' and 
                            existing_lock.get('process_id') == os.getpid()):
                            self.state.is_file_locked = True
                            self.state.lock_holder = 'Dugal Inventory System'
                            self.state.lock_time = datetime.now()
                            return True

                        # Check if lock is stale
                        lock_time = datetime.fromisoformat(existing_lock.get('timestamp', ''))
                        if (datetime.now() - lock_time).total_seconds() > 3600:
                            self.logger.warning("Removing stale lock file")
                            if self.logging_manager:
                                self.logging_manager.log_pattern_match({
                                    'type': 'stale_lock_removed',
                                    'lock_info': existing_lock,
                                    'timestamp': datetime.now().isoformat()
                                })
                            os.remove(lock_path)
                        else:
                            self.logger.warning("File is being edited by another process")
                            if self.logging_manager:
                                self.logging_manager.log_error("File locked by another process", {
                                    'context': 'file_lock',
                                    'lock_holder': existing_lock.get('holder'),
                                    'lock_time': existing_lock.get('timestamp')
                                })
                            
                            if attempt < max_retries - 1:
                                self.logger.warning(f"Retrying lock in {retry_delay} seconds (attempt {attempt+1}/{max_retries})")
                                time.sleep(retry_delay)
                                continue
                            else:
                                return False
                            
                    except (IOError, ValueError, SyntaxError) as e:
                        self.logger.warning(f"Corrupted lock file found, removing: {e}")
                        if self.logging_manager:
                            self.logging_manager.log_error("Corrupted lock file", {
                                'context': 'file_lock',
                                'error': str(e)
                            })
                        os.remove(lock_path)

                # Create new lock file
                lock_info = {
                    'timestamp': datetime.now().isoformat(),
                    'process_id': os.getpid(),
                    'holder': 'Dugal Inventory System',
                    'mode': 'read-only' if self.state.read_only else 'edit',
                    'machine': os.environ.get('COMPUTERNAME', 'Unknown'),
                    'user': os.environ.get('USERNAME', 'Unknown')
                }

                # Use atomic write for lock file
                temp_lock = f"{lock_path}.tmp"
                with open(temp_lock, 'w', encoding='utf-8') as f:
                    f.write(str(lock_info))
                os.replace(temp_lock, lock_path)
                
                self.state.is_file_locked = True
                self.state.lock_holder = 'Dugal Inventory System'
                self.state.lock_time = datetime.now()

                if self.logging_manager:
                    self.logging_manager.log_pattern_match({
                        'type': 'file_locked',
                        'lock_info': lock_info,
                        'timestamp': datetime.now().isoformat()
                    })

                self.logger.debug("File locked successfully")
                return True
                
            except Exception as e:
                error_msg = f"Failed to lock file: {e}"
                self.logger.error(error_msg)
                if self.logging_manager:
                    self.logging_manager.log_error(str(e), {
                        'context': 'file_lock',
                        'action': 'lock',
                        'attempt': attempt + 1
                    })
                
                if attempt < max_retries - 1:
                    self.logger.warning(f"Retrying lock in {retry_delay} seconds (attempt {attempt+1}/{max_retries})")
                    time.sleep(retry_delay)
                else:
                    return False

        return False

    def check_file_locks(self, file_path):
        """Check if a specific file is locked by another process."""
        self.logger.debug(f"Checking if file is locked: {file_path}")
        
        try:
            import win32file
            import pywintypes
            import os
            
            if not os.path.exists(file_path):
                self.logger.debug(f"File does not exist: {file_path}")
                return False
                
            try:
                # Try to open the file exclusively
                handle = win32file.CreateFile(
                    file_path,
                    win32file.GENERIC_READ,
                    0,  # No sharing
                    None,
                    win32file.OPEN_EXISTING,
                    win32file.FILE_ATTRIBUTE_NORMAL,
                    None
                )
                # Close the handle if we got it
                win32file.CloseHandle(handle)
                self.logger.debug(f"File is not locked: {file_path}")
                return False
                
            except pywintypes.error as e:
                self.logger.debug(f"File is locked: {file_path}, error: {e}")
                return True
                
        except Exception as e:
            self.logger.error(f"Error checking file lock: {e}")
            return False

    def unlock_file(self) -> bool:
        """Release file lock with enhanced error handling."""
        if not self.state.local_file_path:
            if self.logging_manager:
                self.logging_manager.log_error("No file selected", {
                    'context': 'file_lock',
                    'action': 'unlock'
                })
            self.logger.error("No file selected for unlocking")
            return False

        lock_path = f"{self.state.local_file_path}.lock"

        try:
            # Verify it's our lock before removing
            if os.path.exists(lock_path):
                try:
                    with open(lock_path, 'r', encoding='utf-8') as f:
                        lock_info = eval(f.read())

                    # Only remove if it's our lock
                    if (lock_info.get('holder') == self.state.lock_holder and 
                        lock_info.get('process_id') == os.getpid()):
                        os.remove(lock_path)
                        if self.logging_manager:
                            self.logging_manager.log_pattern_match({
                                'type': 'file_unlocked',
                                'lock_info': lock_info,
                                'timestamp': datetime.now().isoformat()
                            })
                    else:
                        error_msg = "Attempted to remove another process's lock file"
                        self.logger.warning(error_msg)
                        if self.logging_manager:
                            self.logging_manager.log_error(error_msg, {
                                'context': 'file_lock',
                                'action': 'unlock',
                                'lock_holder': lock_info.get('holder'),
                                'current_holder': self.state.lock_holder
                            })
                        return False

                except (IOError, ValueError, SyntaxError) as e:
                    error_msg = f"Removing corrupted lock file: {e}"
                    self.logger.warning(error_msg)
                    if self.logging_manager:
                        self.logging_manager.log_error(error_msg, {
                            'context': 'file_lock',
                            'action': 'unlock',
                            'error': str(e)
                        })
                    os.remove(lock_path)

            # Reset state locks
            self.state.is_file_locked = False
            self.state.lock_time = None
            self.state.lock_holder = None
            
            self._log_operation('unlock', True)
            
            if self.logging_manager:
                self.logging_manager.log_pattern_match({
                    'type': 'unlock_complete',
                    'file': os.path.basename(self.state.local_file_path),
                    'timestamp': datetime.now().isoformat()
                })
            
            self.logger.debug("File unlocked successfully")
            return True

        except Exception as e:
            error_msg = f"Failed to unlock file: {e}"
            self.logger.error(error_msg)
            if self.logging_manager:
                self.logging_manager.log_error(str(e), {
                    'context': 'file_lock',
                    'action': 'unlock',
                    'file': self.state.local_file_path
                })
            self._log_operation('unlock', False, str(e))
            return False

    def is_read_only_accessible(self, file_path: str) -> bool:
        """Check if file can be accessed in read-only mode."""
        try:
            with open(file_path, 'rb') as f:
                f.read(1)
            self.logger.debug("File accessible in read-only mode: %s", file_path)
            return True
        except Exception as e:
            self.logger.error(f"Read-only access check failed: {e}")
            return False

    def get_search_engine(self):
        """Get the search engine from the component manager or local reference."""
        try:
            # Try to use component manager first
            try:
                from component_manager import component_manager
                
                # Get search engine from component manager
                search_engine = component_manager.get_search_engine()
                
                if search_engine:
                    # If found in component manager, update our local reference
                    if hasattr(self, 'search_engine') and self.search_engine is not search_engine:
                        logger.debug(f"Updating local search engine reference from component manager")
                        self.search_engine = search_engine
                    return search_engine
                    
            except ImportError:
                logger.debug("Component manager not available, falling back to legacy method")
            
            # Fall back to legacy method if component manager is not available
            search_engine = GlobalRegistry.get('search_engine')
            
            # If found in registry, update our local reference
            if search_engine:
                if hasattr(self, 'search_engine') and self.search_engine is not search_engine:
                    logger.debug(f"Updating local search engine reference from registry")
                    self.search_engine = search_engine
                return search_engine
                
            # If not in registry but we have one, register it
            if hasattr(self, 'search_engine') and self.search_engine:
                GlobalRegistry.register('search_engine', self.search_engine)
                return self.search_engine
                
            # Last resort: if excel_handler has one
            if hasattr(self, 'excel_handler') and hasattr(self.excel_handler, 'search_engine'):
                search_engine = self.excel_handler.search_engine
                GlobalRegistry.register('search_engine', search_engine)
                return search_engine
                
            logger.warning("No search engine available")
            return None
            
        except Exception as e:
            logger.error(f"Error getting search engine: {e}")
            # Fallback to local reference if available
            return getattr(self, 'search_engine', None)

    def verify_search_engine(self) -> bool:
        """Verify search engine is properly initialized."""
        if not hasattr(self, 'excel_handler'):
            self.logger.error("No Excel handler available")
            return False
            
        if not hasattr(self.excel_handler, 'search_engine'):
            self.logger.error("No search engine available")
            return False
            
        return True

    def ensure_file_available_with_retry(self, file_path, max_retries=3, retry_delay=1.0):
        """Ensure file is available with automatic retry logic."""
        import time
        
        for attempt in range(max_retries):
            try:
                # Check if file is locked
                if self.check_file_locks(file_path):
                    if attempt < max_retries - 1:
                        self.logger.warning(f"File locked, retrying in {retry_delay} seconds (attempt {attempt+1}/{max_retries})")
                        time.sleep(retry_delay)
                        continue
                    else:
                        self.logger.error(f"File still locked after {max_retries} attempts")
                        return False
                
                # If not locked, try to ensure it's available
                if self.ensure_local_copy(file_path):
                    self.logger.debug(f"File available after {attempt+1} attempts")
                    return True
                
                # If ensure_local_copy failed for other reasons
                if attempt < max_retries - 1:
                    self.logger.warning(f"File not available, retrying in {retry_delay} seconds (attempt {attempt+1}/{max_retries})")
                    time.sleep(retry_delay)
                else:
                    self.logger.error(f"File not available after {max_retries} attempts")
                    
            except Exception as e:
                self.logger.error(f"Error ensuring file availability: {e}")
                if attempt < max_retries - 1:
                    self.logger.warning(f"Retrying in {retry_delay} seconds (attempt {attempt+1}/{max_retries})")
                    time.sleep(retry_delay)
                else:
                    self.logger.error(f"Failed after {max_retries} attempts")
                    return False
        
        return False

    def refresh_file(self) -> bool:
        """Refresh local file from OneDrive."""
        if not self.state.local_file_path:
            if self.logging_manager:
                self.logging_manager.log_error("No file selected", {
                    'context': 'file_refresh',
                    'action': 'refresh'
                })
            self.logger.error("No file selected for refresh")
            return False

        try:
            current_time = datetime.now()
            
            # Check refresh interval
            if (self.state.last_sync_time and 
                (current_time - self.state.last_sync_time).total_seconds() < self.state.sync_interval):
                self.logger.debug("Skipping refresh - within sync interval")
                if self.logging_manager:
                    self.logging_manager.log_pattern_match({
                        'type': 'refresh_skipped',
                        'reason': 'within_interval',
                        'last_sync': self.state.last_sync_time.isoformat(),
                        'timestamp': current_time.isoformat()
                    })
                return True

            if not os.path.exists(self.state.local_file_path):
                error_msg = "File no longer exists"
                if self.logging_manager:
                    self.logging_manager.log_error(error_msg, {
                        'context': 'file_refresh',
                        'file': self.state.local_file_path
                    })
                # Try recovery before giving up
                if self._attempt_file_recovery():
                    self.logger.info("Successfully recovered file path")
                    return self.refresh_file()  # Try again with new path
                raise FileNotFoundError(error_msg)

            current_mtime = os.path.getmtime(self.state.local_file_path)
            
            # Check if file has been modified
            if (self.state.status['last_modified'] and 
                current_mtime > self.state.status['last_modified']):
                self.logger.debug("File modified, creating backup before refresh")
                if self.logging_manager:
                    self.logging_manager.log_pattern_match({
                        'type': 'pre_refresh_backup',
                        'file': os.path.basename(self.state.local_file_path),
                        'timestamp': current_time.isoformat()
                    })
                backup_path = self._create_backup('pre_refresh')
                
                # Log backup creation
                if self.logging_manager:
                    self.logging_manager.log_pattern_match({
                        'type': 'backup_created',
                        'original': os.path.basename(self.state.local_file_path),
                        'backup': os.path.basename(backup_path),
                        'timestamp': current_time.isoformat()
                    })

            time.sleep(0.5)  # Simulate sync delay
            
            self.state.last_sync_time = current_time
            self.state.status.update({
                'last_modified': current_mtime,
                'sync_status': 'synced'
            })
            
            self._log_operation('refresh', True)
            
            if self.logging_manager:
                self.logging_manager.log_pattern_match({
                    'type': 'refresh_complete',
                    'file': os.path.basename(self.state.local_file_path),
                    'timestamp': current_time.isoformat(),
                    'status': self.state.status
                })
                
            # Notify search engine of refresh if available
            if self.search_engine:
                try:
                    self.search_engine.handle_file_refresh(self.state.local_file_path)
                except Exception as e:
                    self.logger.warning(f"Error updating search engine after refresh: {e}")
                    if self.logging_manager:
                        self.logging_manager.log_error(str(e), {
                            'context': 'search_engine_refresh',
                            'file': self.state.local_file_path
                        })

            self.logger.debug("File refresh completed successfully")
            return True

        except Exception as e:
            self.state.sync_failures += 1
            error_msg = f"Failed to refresh file: {e}"
            self.logger.error(error_msg)
            if self.logging_manager:
                self.logging_manager.log_error(str(e), {
                    'context': 'file_refresh',
                    'file': self.state.local_file_path,
                    'sync_failures': self.state.sync_failures
                })
            self._log_operation('refresh', False, str(e))
            
            # Attempt recovery
            try:
                self.logger.debug("Attempting recovery after refresh failure")
                # Try to re-establish file connection
                if self._attempt_file_recovery():
                    self.logger.info("Successfully recovered from refresh failure")
                    return self.refresh_file()  # Try again after recovery
            except Exception as recovery_error:
                self.logger.error(f"Recovery attempt failed: {recovery_error}")
                
            return False

    def _save_state(self):
        """Ensure state is persisted across operations."""
        try:
            # Make sure critical attributes are set and log them
            if not hasattr(self.state, 'local_file_path') or not self.state.local_file_path:
                self.logger.warning("State has empty or missing local_file_path")
            else:
                self.logger.debug(f"State local_file_path: {self.state.local_file_path}")
                
            # If we have a dedicated state persistence mechanism, use it
            if hasattr(self, 'state_manager') and self.state_manager:
                self.state_manager.save_state(self.state)
                self.logger.debug("Saved state via state manager")
                
            # Make sure this state is registered in the global registry if needed
            from global_registry import GlobalRegistry
            if hasattr(self, 'register_state') and self.register_state:
                GlobalRegistry.register('onedrive_handler_state', self.state)
                self.logger.debug("Registered state in global registry")
                
            # Update any related components that might need this state
            if hasattr(self, 'excel_handler') and hasattr(self.excel_handler, 'update_onedrive_state'):
                self.excel_handler.update_onedrive_state(self.state)
                self.logger.debug("Updated ExcelHandler with current state")
                
            return True
        except Exception as e:
            self.logger.error(f"Error saving state: {e}")
            return False

    def refresh_read_only(self) -> bool:
        """Refresh file in read-only mode with enhanced logging and error handling."""
        if not self.state.read_only:
            if self.logging_manager:
                self.logging_manager.log_error("Attempted read-only refresh in edit mode", {
                    'context': 'refresh_read_only',
                    'current_mode': 'edit'
                })
            self.logger.error("Attempted to use read-only refresh in edit mode")
            return False
                
        try:
            current_time = datetime.now()
            
            # Verify file accessibility
            if not self.is_read_only_accessible(self.state.local_file_path):
                error_msg = "File is not accessible in read-only mode"
                self.logger.error(error_msg)
                if self.logging_manager:
                    self.logging_manager.log_error(error_msg, {
                        'context': 'refresh_read_only',
                        'file': self.state.local_file_path,
                        'timestamp': current_time.isoformat()
                    })
                return False
            
            # Check refresh interval
            if (self.state.last_sync_time and 
                (current_time - self.state.last_sync_time).total_seconds() < self.state.sync_interval):
                self.logger.debug("Skipping read-only refresh - within sync interval")
                if self.logging_manager:
                    self.logging_manager.log_pattern_match({
                        'type': 'read_only_refresh_skipped',
                        'reason': 'within_interval',
                        'last_sync': self.state.last_sync_time.isoformat(),
                        'timestamp': current_time.isoformat()
                    })
                return True
                    
            # Update state
            self.state.last_sync_time = current_time
            self.state.status.update({
                'sync_status': 'synced',
                'last_modified': os.path.getmtime(self.state.local_file_path),
                'access_count': self.state.status.get('access_count', 0) + 1
            })
            
            if self.logging_manager:
                self.logging_manager.log_pattern_match({
                    'type': 'read_only_refresh_complete',
                    'file': os.path.basename(self.state.local_file_path),
                    'timestamp': current_time.isoformat(),
                    'status': self.state.status
                })
            
            # Notify search engine if available
            if self.search_engine:
                try:
                    self.search_engine.handle_read_only_refresh(self.state.local_file_path)
                except Exception as e:
                    self.logger.warning(f"Error updating search engine after read-only refresh: {e}")
                    if self.logging_manager:
                        self.logging_manager.log_error(str(e), {
                            'context': 'search_engine_read_only_refresh',
                            'file': self.state.local_file_path
                        })

            self.logger.debug("Read-only refresh completed successfully")
            return True
                
        except Exception as e:
            error_msg = f"Read-only refresh failed: {e}"
            self.logger.error(error_msg)
            if self.logging_manager:
                self.logging_manager.log_error(str(e), {
                    'context': 'refresh_read_only',
                    'file': self.state.local_file_path,
                    'status': self.state.status
                })
            self.state.sync_failures += 1
            return False

    def update_inventory(self, item_name: str, value: float, is_addition: bool = False) -> Dict[str, Any]:
        """Update inventory for the specified item."""
        try:
            # First, ensure the file is locked for editing
            if not self.state.is_file_locked:
                self.logger.debug(f"File not locked for updates, attempting to lock now: {self.state.local_file_path}")
                locked = self.lock_file()
                if not locked:
                    error_msg = "Failed to lock file for updates"
                    if self.logging_manager:
                        self.logging_manager.log_error(error_msg, {
                            'context': 'inventory_update',
                            'item': item_name
                        })
                    self.logger.error(error_msg)
                    return {"success": False, "message": error_msg}
                self.logger.debug("Successfully locked file for inventory update")
            
            # At this point, the file should be locked - double check
            if not self.state.is_file_locked:
                error_msg = "File still not locked for updates despite locking attempt"
                if self.logging_manager:
                    self.logging_manager.log_error(error_msg, {
                        'context': 'inventory_update',
                        'item': item_name
                    })
                self.logger.error(error_msg)
                return {"success": False, "message": error_msg}

            # Now continue with the original update logic
            self.logger.debug(f"Attempting to update inventory for {item_name} with value {value}, operation: {'add/subtract' if is_addition else 'set'}")
            if self.logging_manager:
                self.logging_manager.log_pattern_match({
                    'type': 'inventory_update_start',
                    'item': item_name,
                    'value': value,
                    'operation': 'add/subtract' if is_addition else 'set',
                    'timestamp': datetime.now().isoformat()
                })

            # Get the search engine using the standard accessor if available
            search_engine = self.get_search_engine() if hasattr(self, 'get_search_engine') else self.search_engine

            workbook = load_workbook(self.state.local_file_path, data_only=True)
            updates = []

            # Verify required handlers
            if not (hasattr(self, 'excel_handler') or search_engine):
                error_msg = "No search system available"
                if self.logging_manager:
                    self.logging_manager.log_error(error_msg, {
                        'context': 'inventory_update',
                        'search_systems': {
                            'excel_handler': hasattr(self, 'excel_handler'),
                            'search_engine': bool(search_engine)
                        }
                    })
                self.logger.error(error_msg)
                return {"success": False, "message": error_msg}

            for sheet_name in self.state.selected_sheets:
                self.logger.debug(f"Checking sheet {sheet_name}")
                sheet = workbook[sheet_name]

                try:
                    search_col_idx = column_index_from_string(self.state.search_column)
                    self.logger.debug(f"Using search column {self.state.search_column} (index: {search_col_idx})")

                    # Try search engine first, fall back to excel_handler
                    match_row = None
                    if self.search_engine:
                        match_row = self.search_engine.search(sheet, item_name)
                        if match_row and self.logging_manager:
                            self.logging_manager.log_pattern_match({
                                'type': 'search_engine_match',
                                'item': item_name,
                                'sheet': sheet_name,
                                'row': match_row
                            })
                    
                    # Fallback to excel_handler if no match found
                    if not match_row and hasattr(self, 'excel_handler'):
                        match_row = self.excel_handler.search_inventory(
                            sheet=sheet,
                            search_term=item_name,
                            column_index=search_col_idx
                        )

                    if match_row:
                        self.logger.debug(f"Found match in row {match_row} of sheet {sheet_name}")
                        # Process match...
                        try:
                            input_col = self.find_column_by_header(self.state.input_column, sheet)
                            if not input_col:
                                error_msg = f"Could not find input column '{self.state.input_column}' in sheet '{sheet_name}'"
                                self.logger.error(error_msg)
                                if self.logging_manager:
                                    self.logging_manager.log_error(error_msg, {
                                        'context': 'column_search',
                                        'sheet': sheet_name
                                    })
                                continue

                            self.logger.debug(f"Using input column '{self.state.input_column}' (index: {input_col})")
                            current_value = sheet.cell(row=match_row, column=input_col).value
                            
                            try:
                                current_value = float(current_value or 0)
                                
                                # Modified logic to handle add/subtract vs. direct set
                                if is_addition:
                                    # Add/subtract the value to/from current value
                                    new_value = current_value + value
                                    self.logger.debug(f"Adding {value} to current value {current_value} = {new_value}")
                                else:
                                    # Direct replacement
                                    new_value = value
                                    self.logger.debug(f"Setting value to {value} (was {current_value})")
                                
                                # Update the cell
                                sheet.cell(row=match_row, column=input_col).value = new_value
                                self.logger.debug(f"Updated cell value in sheet {sheet_name}, row {match_row}, col {input_col}: {current_value} -> {new_value}")
                                
                                # Cache the update
                                cache_key = f"{sheet_name}_{match_row}_{input_col}"
                                self.state.cache['values'][cache_key] = new_value
                                
                                # Get the actual item name from the sheet for learning
                                actual_item = str(sheet.cell(row=match_row, column=search_col_idx).value)
                                
                                updates.append({
                                    "sheet": sheet_name,
                                    "item": actual_item,
                                    "old_value": current_value,
                                    "new_value": new_value,
                                    "operation": "add/subtract" if is_addition else "set",
                                    "row": match_row,
                                    "column": input_col
                                })
                                
                                # Record successful match for learning
                                if self.search_engine:
                                    try:
                                        self.search_engine.record_successful_match(
                                            search_term=item_name,
                                            matched_item=actual_item
                                        )
                                        self.logger.debug(f"Recorded match pattern: {item_name} -> {actual_item}")
                                    except Exception as learning_error:
                                        self.logger.error(f"Error recording match pattern: {learning_error}")
                                        if self.logging_manager:
                                            self.logging_manager.log_error(str(learning_error), {
                                                'context': 'pattern_learning',
                                                'search_term': item_name,
                                                'matched_item': actual_item
                                            })
                                
                            except ValueError as e:
                                error_msg = f"Value error updating {item_name} in sheet {sheet_name}: {e}"
                                self.logger.error(error_msg)
                                if self.logging_manager:
                                    self.logging_manager.log_error(str(e), {
                                        'context': 'value_conversion',
                                        'item': item_name,
                                        'sheet': sheet_name
                                    })
                                continue
                        except Exception as e:
                            error_msg = f"Error processing match in sheet {sheet_name}: {e}"
                            self.logger.error(error_msg)
                            if self.logging_manager:
                                self.logging_manager.log_error(str(e), {
                                    'context': 'match_processing',
                                    'sheet': sheet_name
                                })
                            continue
                    else:
                        self.logger.debug(f"No match found in sheet {sheet_name}")

                except ValueError as e:
                    error_msg = f"Invalid search column: {self.state.search_column}"
                    self.logger.error(error_msg)
                    self.logger.debug(f"ValueError details: {str(e)}")
                    if self.logging_manager:
                        self.logging_manager.log_error(str(e), {
                            'context': 'column_validation',
                            'column': self.state.search_column
                        })
                    continue
                
            if updates:
                # Save the workbook
                workbook.save(self.state.local_file_path)
                self.file_updated.emit()
                operation_type = "added to/subtracted from" if is_addition else "set"
                success_msg = f"{operation_type} {item_name} in {len(updates)} sheets"
                self._log_operation('update', True, success_msg)
                if self.logging_manager:
                    self.logging_manager.log_pattern_match({
                        'type': 'update_success',
                        'item': item_name,
                        'operation': 'add/subtract' if is_addition else 'set',
                        'updates': updates,
                        'timestamp': datetime.now().isoformat()
                    })
                return {
                    "success": True,
                    "updates": updates,
                    "message": success_msg
                }
            
            error_msg = f"No matching items found for {item_name}"
            self.logger.debug(error_msg)
            if self.logging_manager:
                self.logging_manager.log_pattern_match({
                    'type': 'update_no_matches',
                    'item': item_name,
                    'timestamp': datetime.now().isoformat()
                })
            return {
                "success": False,
                "message": error_msg
            }
            
        except Exception as e:
            error_msg = f"Error updating inventory: {str(e)}"
            self.logger.error(error_msg)
            self._log_operation('update', False, str(e))
            if self.logging_manager:
                self.logging_manager.log_error(str(e), {
                    'context': 'inventory_update',
                    'item': item_name,
                    'value': value,
                    'operation': 'add/subtract' if is_addition else 'set'
                })
            return {
                "success": False,
                "message": error_msg
            }

    def find_column_by_header(self, header: str, sheet) -> Optional[int]:
        """
        Find the column index for a given header in the first row of a sheet.
        
        Args:
            header: The header name to search for.
            sheet: The worksheet to search.
            
        Returns:
            Optional[int]: The column index (1-based) of the header, or None if not found.
        """
        cache_key = f"{sheet.title}_{header.strip().lower()}"
        # Check cache first
        if cache_key in self.state.cache['columns']:
            self.logger.debug(f"Cache hit for header '{header}' in sheet '{sheet.title}': Column {self.state.cache['columns'][cache_key]}")
            return self.state.cache['columns'][cache_key]
        
        try:
            # Iterate through the first row to find the header
            for idx, cell in enumerate(sheet[1], start=1):  # sheet[1] is the first row
                if cell.value:
                    cell_value = str(cell.value).strip().lower()
                    if cell_value == header.strip().lower():
                        self.state.cache['columns'][cache_key] = idx
                        self.logger.debug(f"Header '{header}' found in sheet '{sheet.title}' at column {idx}")
                        return idx
            
            self.logger.debug(f"Header '{header}' not found in sheet '{sheet.title}'")
            return None  # Header not found

        except Exception as e:
            self.logger.error(f"Error finding header '{header}' in sheet '{sheet.title}': {e}")
            return None

    def _create_backup(self, backup_type: str) -> str:
        """Create a backup of the current file."""
        if not self.state.local_file_path or not os.path.exists(self.state.local_file_path):
            raise FileNotFoundError("No file to backup")

        timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
        backup_name = f"{os.path.basename(self.state.local_file_path)}.{backup_type}_{timestamp}"
        backup_path = os.path.join(self.backup_dir, backup_name)
        
        shutil.copy2(self.state.local_file_path, backup_path)
        self.state.status['last_backup'] = datetime.now()
        
        # Update operations history
        self._log_operation('backup', True, f"Created {backup_type} backup")
        
        return backup_path

    def _cleanup_old_backups(self) -> None:
        """Clean up old backup files, keeping only the most recent 5."""
        try:
            if not self.state.local_file_path:
                return

            backups = sorted(
                [f for f in os.listdir(self.backup_dir) 
                 if f.startswith(os.path.basename(self.state.local_file_path))],
                key=lambda x: os.path.getmtime(os.path.join(self.backup_dir, x))
            )

            # Remove all but the 5 most recent backups
            for backup in backups[:-5]:
                try:
                    os.remove(os.path.join(self.backup_dir, backup))
                except OSError as e:
                    self.logger.error(f"Error removing backup {backup}: {e}")

        except Exception as e:
            self.logger.error(f"Error cleaning up old backups: {e}")

    def _verify_content_hash(self, current_content: bytes, backup_path: str) -> bool:
        """Verify file hasn't been modified by comparing with backup."""
        with open(backup_path, 'rb') as f:
            backup_content = f.read()
        return current_content == backup_content

    def _log_operation(self, operation_type: str, success: bool, details: Optional[str] = None) -> None:
        """Log operations for tracking and debugging."""
        operation = {
            'timestamp': datetime.now().isoformat(),
            'type': operation_type,
            'success': success,
            'details': details
        }
        
        self.state.recent_operations.append(operation)
        if len(self.state.recent_operations) > self.max_operations_history:
            self.state.recent_operations.pop(0)

        if not success:
            self.state.status['error_count'] += 1

    def _get_dugal_response(self, category: str, context: Optional[str] = None) -> str:
        """Get a response from Dugal if available."""
        if self.state.dugal and hasattr(self.state.dugal, 'get_response'):
            return self.state.dugal.get_response(category, context)
        return context or category

    def _is_lock_stale(self, lock_path: str) -> bool:
        """Check if a lock file is stale (older than 4 hours)."""
        try:
            lock_time = datetime.fromtimestamp(os.path.getmtime(lock_path))
            return (datetime.now() - lock_time).total_seconds() > 14400  # 4 hours
        except Exception as e:
            self.logger.error(f"Error checking lock staleness: {e}")
            return True  # Assume stale if we can't check

    def cleanup(self) -> None:
        """Clean up resources and remove locks before shutdown."""
        try:
            if self.state.is_file_locked:
                self.unlock_file()

            # Clean up backup directory
            self._cleanup_old_backups()
            
            # Clear caches
            self.state.cache = {
                'items': {},
                'values': {},
                'matches': {}
            }
            
            # Reset state
            self.state = OneDriveState()
            
            self.logger.info("OneDrive handler cleanup completed")
        except Exception as e:
            self.logger.error(f"Error during cleanup: {e}")
            
    def _attempt_file_recovery(self) -> bool:
        """Attempt to recover file access after a failure."""
        try:
            self.logger.debug("Attempting file recovery")
            
            # Check if file exists
            if not os.path.exists(self.state.local_file_path):
                self.logger.warning("File not found, attempting to locate")
                # Try to find the file in common locations
                common_locations = [
                    os.path.join(os.path.expanduser("~"), "Desktop"),
                    os.path.join(os.path.expanduser("~"), "OneDrive", "Desktop"),
                    os.path.join(os.path.expanduser("~"), "Documents")
                ]
                
                filename = os.path.basename(self.state.local_file_path)
                for location in common_locations:
                    potential_path = os.path.join(location, filename)
                    if os.path.exists(potential_path):
                        self.logger.info(f"Found file at {potential_path}")
                        self.state.local_file_path = potential_path
                        return True
                        
                return False
                
            # Reset sync state
            self.state.last_sync_time = None
            self.state.sync_failures = 0
            
            return True
            
        except Exception as e:
            self.logger.error(f"Error in file recovery: {e}")
            return False
